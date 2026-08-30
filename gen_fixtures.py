#!/usr/bin/env python3
"""Regenerate every test fixture under testdata/.

Fixtures are generated, never hand-edited: a fixture whose provenance is "I
tweaked a byte once" cannot be reasoned about when a test starts failing.
Each generator under testdata/gen/ owns a disjoint set of files and documents,
in its own docstring, what each file stresses and what a correct parse must
produce.

    python3 gen_fixtures.py              # everything
    python3 gen_fixtures.py 01 04        # only the generators whose name matches
    python3 gen_fixtures.py --list

Requires openpyxl for the spreadsheet fixtures, and xlwt for the legacy .xls
ones (09 skips those with a notice if it is missing). ODS is written with the
standard library.
"""

import os
import re
import subprocess
import sys
import zipfile
from datetime import datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
GEN_DIR = os.path.join(ROOT, "testdata", "gen")


def generators():
    if not os.path.isdir(GEN_DIR):
        return []
    return sorted(
        f for f in os.listdir(GEN_DIR) if f.endswith(".py") and not f.startswith("_")
    )


_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _repack_deterministic(path):
    """Rewrite an xlsx zip with pinned entry timestamps and dcterms:modified."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        entries = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    entries = [
        (name, _MODIFIED_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", data)
         if name == "docProps/core.xml" else data)
        for name, data in entries
    ]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for name, data in entries:
            info = zipfile.ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            zout.writestr(info, data)
    os.replace(tmp, path)


def umsatz():
    """The canonical nightmare file: title block, two-row header with German
    months, vertically merged Region cells, an interleaved subtotal row, and a
    Total footer. Referenced by name from tests/e2e.rs."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Umsatz"

    # 3-row title block
    ws.append(["Muster AG — Umsatzübersicht"])
    ws.append(["Erstellt: 05.01.2026, Abteilung Controlling"])
    ws.append([])

    # Two-row header: year on top (merged across months), months below
    ws.append(["Region", "Produkt", 2025, None, None, None])
    ws.append([None, None, "Jan", "Feb", "Mär", "Dez"])
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=6)
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)

    # Body: Region vertically merged (value only in top cell), one subtotal row
    rows = [
        ["Ost", "Widget", 1200.50, 1300.00, 990.25, 1500.75],
        [None, "Gadget", 800.00, 850.50, 700.00, 1000.00],
        ["Zwischensumme", None, 2000.50, 2150.50, 1690.25, 2500.75],
        ["West", "Widget", 2100.00, 2200.25, 1800.00, 2600.50],
        [None, "Gadget", 950.75, 900.00, 1100.50, 1250.25],
    ]
    for r in rows:
        ws.append(r)
    ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1)
    ws.merge_cells(start_row=9, start_column=1, end_row=10, end_column=1)

    # Footer
    ws.append(["Total", None, 5051.25, 5250.75, 4590.75, 6351.50])

    out = os.path.join(ROOT, "testdata", "umsatz.xlsx")
    # Pin the document properties and repack the zip with fixed entry stamps,
    # or this file's bytes change on every run and stale the fingerprint in
    # every committed sidecar that points at it. openpyxl rewrites
    # dcterms:modified at save time whatever the properties say, so the XML
    # is patched as well.
    wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
    wb.save(out)
    _repack_deterministic(out)
    print(f"wrote {os.path.relpath(out, ROOT)}")


def main() -> int:
    args = [a for a in sys.argv[1:]]
    if "--list" in args:
        print("umsatz (built in)")
        for g in generators():
            print(g)
        return 0

    os.makedirs(os.path.join(ROOT, "testdata"), exist_ok=True)
    selected = [g for g in generators() if not args or any(a in g for a in args)]

    failures = []
    if not args or any(a in "umsatz" for a in args):
        try:
            umsatz()
        except Exception as e:  # noqa: BLE001 - report and keep going
            failures.append(("umsatz", e))

    for g in selected:
        print(f"--- {g}")
        r = subprocess.run(
            [sys.executable, os.path.join("testdata", "gen", g)], cwd=ROOT
        )
        if r.returncode != 0:
            failures.append((g, f"exit {r.returncode}"))

    if failures:
        print("\nFAILED:", file=sys.stderr)
        for name, err in failures:
            print(f"  {name}: {err}", file=sys.stderr)
        return 1
    print("\nall fixtures regenerated")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
