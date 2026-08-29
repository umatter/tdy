#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fixture generator: Swiss/German financial exports (job key: swiss-german-finance).

Run from the repo root:   python3 testdata/gen/01_swiss_finance.py
Deterministic and idempotent: reruns produce byte-identical files (the .xlsx zip
is rewritten with fixed entry timestamps and fixed docProps), so blake3 sidecar
fingerprints stay stable across regenerations.

Every file it writes lives in testdata/swiss_german_finance/ and is named
swiss_german_finance_*, so it can never collide with another generator's output.

WHAT EACH FIXTURE STRESSES / WHAT A CORRECT PARSE MUST PRODUCE
=============================================================

All row/column claims below describe the *correct* parse, i.e. what a
hand-written or LLM-repaired ParseSpec must yield. Where tier-1 (src/sniff.rs)
is known to produce an executable-but-wrong spec, that is called out; those
cases are the point of the fixture.

01_kontoauszug_utf8.csv  (UTF-8 **with BOM**, LF, ';'-delimited)
    Flagship Swiss bank statement. 2 title rows of different arity, blank line,
    German header row with umlauts/spaces/punctuation, 12 data rows, one
    interleaved "Zwischentotal" row, one "Total" footer row.
    Correct parse -> 12 rows x 8 columns:
      buchung        date   %d.%m.%Y      first value 2026-01-05
      valuta         date   %d.%m.%y      first value 2026-01-07 (buchung + 2d)
      text           utf8                 row 3 = "Rückzahlung; Teilbetrag"
      betrag_chf     decimal(12,2)        "CHF 1'200.00" -> 1200.00,
                                          "(CHF 1'234.55)" -> -1234.55, 3 nulls,
                                          SUM = 1269728.82
      betrag_eur     decimal(12,2)        "1.234,56 EUR" -> 1234.56, 4 nulls,
                                          SUM = 1517761.18
      waehrungskurs  float64              "1,5" -> 1.5 (NOT 15), SUM = 16.5
      quote          float64              "12,5%" -> 12.5, 2 nulls, SUM = 125.1
      beleg_nr       int64                "4001*" -> 4001, SUM = 48078
    Tier-1 traps (all silent-corruption class, invariant 6):
      * "1,5" -> the thousands-separator candidate (',', None) fires first and
        yields 15.0; the whole waehrungskurs column comes out 10x-100x too big.
      * "05.01.2026" -> the ('.', ',') candidate yields 5012026.0, so BOTH date
        columns are typed float64 before the date formats are ever tried.
      * "1.234,56 EUR" / "CHF 1'200.00" / "12,5%" / "4001*" are safely left utf8
        (no corruption) because the affix blocks every numeric candidate.
    Tier-1 ships all of that as method="heuristic" at confidence 0.90 (0.95 minus
    0.05 for the skipped title rows), i.e. >= the 0.8 threshold, so nothing ever
    escalates to the LLM tier. Its structural guess is skip_rows head=2 (the
    blank line is not a CSV record at all) + promote_header rows=1, with no
    footer handling: 14 body rows instead of 12.

02_kontoauszug_latin1.csv  (windows-1252, LF, ';'-delimited, no BOM)
    Byte-different, semantically identical twin of fixture 01 (same 12 rows,
    same sums). Contains ü ä ö é, en dash U+2013 (0x96) and em dash U+2014
    (0x97) - the two bytes that separate windows-1252 from ISO-8859-1.
    Correct parse: identical RecordBatch to fixture 01 once the spec declares
    encoding = "windows-1252". Decoding it as UTF-8 does not error - it yields
    U+FFFD replacement chars, i.e. silently mangled text.

03_dezimalkomma.csv  (UTF-8, ';'-delimited, no title block)
    The classic mis-parse, isolated and unambiguous: two columns of German
    decimal commas with NO thousands separator.
    Correct parse -> 6 rows x 4 columns: artikel utf8, menge_kg float64,
    preis_chf float64 (both decimal_separator = ','), lager utf8.
    menge_kg SUM = 27.125, preis_chf SUM = 18.75.
    Tier-1 yields exactly 10x/100x too much (SUM 1745 and 1875) with
    thousands_separator = ',' - structurally right, numerically wrong, and it
    ships at confidence 0.95 with no warning at all.

04_bankauszug.xlsx  (2 sheets, "Kontoauszug" first)
    Excel twin of fixture 01 with the Excel-specific hazards layered on:
    3-row title block, TWO-row header with a horizontally merged "Betrag"
    group cell, vertically merged account cells, one "Zwischentotal" row, one
    "Total" footer, a second "Legende" sheet of prose.
    Money columns are stored as TEXT (as SAP/Abacus exports do) while Kurs is a
    REAL float and Datum is a REAL Excel date.
    Correct parse -> 12 rows x 8 columns: konto utf8 (fill_down), datum
    timestamp %Y-%m-%d %H:%M:%S (calamine renders date cells WITH a 00:00:00
    time part), kurs float64, quote float64, beleg int64, status utf8,
    betrag_chf decimal(12,2) SUM = 1269728.82, betrag_eur decimal(12,2)
    SUM = 1517761.18 - the same two sums as fixture 01.
    Traps: (a) a real Excel error cell (t="e") in `status` reaches the engine as
    calamine's "#ERR:..." rendering, NOT as the "#N/A" a human sees in Excel, so
    na_values = ["#N/A"] does not null it; (b) a *different* cell holds the
    literal string "#N/A" (t="inlineStr") which na_values DOES null; (c) the
    float cell 1.0 renders as "1", not "1.0"; (d) tier-1's Excel header
    heuristic picks sheet row 6 (the FIRST DATA ROW) as the header, because the
    second header row is too sparse to satisfy its `next_ok` test - so booking 1
    is consumed as column names and only 11 rows survive. Confidence lands at
    exactly 0.80 and the escalation test is `>=`, so that spec ships.

05_grenzfaelle.csv  (UTF-8, ';'-delimited)
    One row per pathology, 6 rows x 8 columns, each numeric column internally
    consistent so a single ValueParsing can express it:
      fall            utf8         "A".."F"
      chf_apostroph   decimal(14,2) sign before AND after the currency word
                                   ("CHF -50.00" and "-CHF 50.00"), SUM = 12345344.34
      chf_strich      decimal(12,2) Swiss cents shorthand "1'234.–" (EN DASH,
                                   U+2013), needs replace ".–" -> "",
                                   SUM = 1013326.00
      eur_schmalraum  decimal(12,2) NBSP (U+00A0) as thousands separator,
                                   SUM = 1011734.06
      klammer_negativ decimal(12,2) parentheses negatives, needs replace
                                   "(" -> "-" and ")" -> "", SUM = -1012580.24
      prozent         float64      "12,5 %" with a space before the sign, SUM = 145.33
      fussnote        int64        "1235 *" / "1238 ***", SUM = 7419
      fehlwerte       decimal(12,2) 5 nulls ("n/a", "-", "k.A.", "#N/A", "") and
                                   one value 1234.00
    The parentheses column is the silent-corruption canary: stripping "[()]"
    instead of replacing yields +1234.56 for "(1.234,56)", turning the correct
    SUM of -1012580.24 into +1014580.24 - plausible magnitudes, flipped signs,
    no error raised anywhere.
    Tier-1 types all 8 columns utf8 (safe but useless) at confidence exactly
    0.80 after the all-utf8 penalty, and never recognises "k.A." as a null:
    it is absent from the sniffer's NA_TOKENS list.

06_gemischte_konvention.csv  (UTF-8, ';'-delimited)
    A single money column with irreconcilably mixed locales (Swiss, German,
    plain, Anglo). No (thousands, decimal) pair parses all six rows.
    Correct behaviour is to REFUSE: either leave `betrag` as utf8 (what tier-1
    does) or fail loudly. With thousands="'" / decimal="." the engine must
    error with `row 2: cannot parse "1.234,56"` - never emit a number.
    -> 6 rows x 3 columns: id int64, betrag utf8, quelle utf8.
    Tier-1 gets this one right (betrag stays utf8, confidence 0.95): no
    separator candidate parses all six rows, so it declines to type the column.
"""

import csv
import datetime
import io
import os
import random
import re
import zipfile

from openpyxl import Workbook
from openpyxl.cell.cell import ERROR_CODES

# No randomness is used below (every value is pinned), but the seed is set so
# that any future filler stays reproducible.
random.seed(20260828)

PREFIX = "swiss_german_finance"
REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTDIR = os.path.join(REPO, "testdata", PREFIX)

NBSP = " "   # U+00A0 no-break space, cp1252 0xA0
ENDASH = "–"  # cp1252 0x96
EMDASH = "—"  # cp1252 0x97


def out(name):
    return os.path.join(OUTDIR, "%s_%s" % (PREFIX, name))


def written(path):
    print("wrote %s (%d bytes)" % (os.path.relpath(path, REPO), os.path.getsize(path)))


def csv_text(rows, delimiter=";"):
    """Render rows to CSV text with LF terminators and minimal quoting."""
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL,
                   lineterminator="\n")
    for r in rows:
        w.writerow(r)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Shared statement body: fixtures 01, 02 and 04 all describe the same 12
# bookings, so their CHF/EUR/kurs/quote/beleg sums must agree.
# ---------------------------------------------------------------------------

# (buchung, text, betrag_chf, betrag_eur, kurs, quote, beleg, status)
BOOKINGS = [
    (datetime.date(2026, 1, 5),  "Lohnzahlung Januar",
     "CHF 1'200.00", "1.234,56 EUR", "1,08", "12,5%", "4001*", "gebucht"),
    (datetime.date(2026, 1, 12), "Bürobedarf Papeterie Müller",
     "CHF 85.50", "987,50 EUR", "1,5", "0,75%", "4002", "gebucht"),
    (datetime.date(2026, 1, 19), "Rückzahlung; Teilbetrag",
     "(CHF 1'234.55)", "k.A.", "0,97", "-1,25%", "4003", "storniert"),
    (datetime.date(2026, 1, 26), "Miete März, Büro Zürich-Höngg",
     "CHF 12'000.00", "12.000,00 EUR", "2,75", "n/a", "4004**", "gebucht"),
    (datetime.date(2026, 2, 2),  "Zinsgutschrift",
     "n/a", "-45,20 EUR", "1,00", "0,00%", "4005", "gebucht"),
    (datetime.date(2026, 2, 9),  "Spesen Zürich%sGenf" % ENDASH,
     "CHF 0.05", "0,99 EUR", "0,05", "100,00%", "4006", "gebucht"),
    # --- Zwischentotal is inserted here ---
    (datetime.date(2026, 2, 16), 'Gebühren "Paket Gold"',
     "(CHF 99.95)", "#N/A", "1,25", "3,5%", "4007*", "ERRORCELL"),
    (datetime.date(2026, 2, 23), "Überweisung Grossmünster AG",
     "CHF 250'000.00", "1.500.000,00 EUR", "3,4", "-0,5%", "4008", "gebucht"),
    (datetime.date(2026, 3, 2),  "Rückerstattung MwSt.",
     "-", "250,00 EUR", "0,5", "k.A.", "4009", "pendent"),
    (datetime.date(2026, 3, 9),  "Dividende Q4 (Société Générale)",
     "CHF 7'777.77", "-", "1,125", "2,25%", "4010", "gebucht"),
    (datetime.date(2026, 3, 16), "Umbuchung Fremdwährung",
     "k.A.", "n/a", "2,0", "0,1%", "4011***", "gebucht"),
    (datetime.date(2026, 3, 31), "Abschluss Quartal",
     "CHF 1'000'000.00", "3.333,33 EUR", "0,875", "7,75%", "4012", "gebucht"),
]

SUBTOTAL_AFTER = 6  # rows 1..6 are summed by the Zwischentotal row
SUB_CHF = "CHF 12'051.00"
SUB_EUR = "14.177,85 EUR"
TOT_CHF = "CHF 1'269'728.82"
TOT_EUR = "1.517.761,18 EUR"


def statement_rows():
    """The CSV shape of the shared statement: header + body + subtotal + total."""
    header = ["Buchung", "Valuta", "Text", "Betrag CHF", "Betrag EUR",
              "Währungskurs", "Quote %", "Beleg-Nr."]
    rows = [header]
    for i, (d, text, chf, eur, kurs, quote, beleg, _status) in enumerate(BOOKINGS):
        valuta = d + datetime.timedelta(days=2)
        rows.append([
            d.strftime("%d.%m.%Y"),
            valuta.strftime("%d.%m.%y"),
            text,
            chf, eur, kurs, quote, beleg,
        ])
        if i + 1 == SUBTOTAL_AFTER:
            rows.append(["", "", "Zwischentotal Jan-Feb", SUB_CHF, SUB_EUR, "", "", ""])
    rows.append(["", "", "Total", TOT_CHF, TOT_EUR, "", "", ""])
    return rows


def statement_document():
    """Title block (ragged arity) + blank line + the tabular part."""
    title1 = "Muster Privatbank AG %s Kontoauszug Q1 2026\n" % EMDASH
    title2 = "Konto 12-345678-9;Währung: CHF/EUR;Stand 31.03.2026\n"
    return title1 + title2 + "\n" + csv_text(statement_rows())


# ---------------------------------------------------------------------------
# 01 + 02: the same statement, UTF-8-with-BOM and windows-1252
# ---------------------------------------------------------------------------

def write_statement_csvs():
    doc = statement_document()

    p1 = out("01_kontoauszug_utf8.csv")
    with open(p1, "wb") as fh:
        fh.write(b"\xef\xbb\xbf")          # UTF-8 BOM, as "CSV UTF-8" exports have
        fh.write(doc.encode("utf-8"))
    written(p1)

    p2 = out("02_kontoauszug_latin1.csv")
    with open(p2, "wb") as fh:
        fh.write(doc.encode("cp1252"))     # no BOM; 1 byte per character
    written(p2)


# ---------------------------------------------------------------------------
# 03: German decimal commas, no thousands separator
# ---------------------------------------------------------------------------

def write_dezimalkomma():
    rows = [
        ["Artikel", "Menge kg", "Preis CHF", "Lager"],
        ["Mehl", "1,5", "2,40", "A"],
        ["Zucker", "2,75", "1,95", "B"],
        ["Salz", "0,5", "0,80", "A"],
        ["Hefe", "12,25", "0,45", "C"],
        ["Butter", "0,125", "9,90", "B"],
        ["Öl", "10,0", "3,25", "A"],
    ]
    p = out("03_dezimalkomma.csv")
    with open(p, "wb") as fh:
        fh.write(csv_text(rows).encode("utf-8"))
    written(p)


# ---------------------------------------------------------------------------
# 04: the Excel twin
# ---------------------------------------------------------------------------

def text_cell(ws, row, col, value):
    """Write a *string* cell, even for values openpyxl would type as an Excel
    error (#N/A and friends land in ERROR_CODES and become t="e" otherwise)."""
    c = ws.cell(row=row, column=col, value=value)
    if isinstance(value, str) and value in ERROR_CODES:
        c.data_type = "s"
    return c


def write_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontoauszug"

    # rows 1-3: title block (row 3 deliberately blank)
    ws.append(["Muster Privatbank AG %s Kontoauszug Q1 2026" % EMDASH])
    ws.append(["Konto 12-345678-9 %s Währung CHF/EUR %s Stand 31.03.2026"
               % (ENDASH, ENDASH)])
    ws.append([])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # rows 4-5: two-row header, "Betrag" merged horizontally over G:H
    ws.append(["Konto", "Datum", "Kurs", "Quote %", "Beleg-Nr.", "Status",
               "Betrag", None])
    ws.append([None, None, None, None, None, None, "CHF", "EUR"])
    for col in range(1, 7):                       # A..F merged vertically
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=8)

    konto_a = "CH93 0076 2011 6238 5295 7"
    konto_b = "CH56 0483 5012 3456 7800 9"

    row = 6
    group_start = {}
    for i, (d, _text, chf, eur, kurs, quote, beleg, status) in enumerate(BOOKINGS):
        konto = None
        if i == 0:
            konto = konto_a
            group_start["a"] = row
        elif i == SUBTOTAL_AFTER:
            konto = konto_b
            group_start["b"] = row + 1  # the Zwischentotal row is written first

        if i == SUBTOTAL_AFTER:
            ws.cell(row=row, column=1, value="Zwischentotal Jan-Feb")
            ws.cell(row=row, column=7, value=SUB_CHF)
            ws.cell(row=row, column=8, value=SUB_EUR)
            row += 1

        ws.cell(row=row, column=1, value=konto)        # blank = vertically merged
        ws.cell(row=row, column=2, value=d)            # REAL Excel date
        ws.cell(row=row, column=3, value=float(kurs.replace(",", ".")))  # REAL float
        ws.cell(row=row, column=4, value=quote)        # text "12,5%"
        ws.cell(row=row, column=5, value=beleg)        # text "4001*"

        if status == "ERRORCELL":
            # openpyxl types strings in ERROR_CODES as t="e": a genuine Excel
            # error cell, which calamine surfaces as "#ERR:...", not "#N/A".
            ws.cell(row=row, column=6, value="#N/A")
        else:
            ws.cell(row=row, column=6, value=status)

        # The money columns keep "#N/A" as a literal *string* cell (t="inlineStr");
        # only `status` above carries a genuine Excel error cell. The pair is the
        # point: na_values = ["#N/A"] nulls the string one and misses the error one.
        text_cell(ws, row, 7, chf)
        text_cell(ws, row, 8, eur)
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=7, value=TOT_CHF)
    ws.cell(row=total_row, column=8, value=TOT_EUR)

    # Vertical merges for the two account blocks (value in the top cell only).
    ws.merge_cells(start_row=group_start["a"], start_column=1,
                   end_row=group_start["a"] + SUBTOTAL_AFTER - 1, end_column=1)
    ws.merge_cells(start_row=group_start["b"], start_column=1,
                   end_row=total_row - 1, end_column=1)

    legende = wb.create_sheet("Legende")
    for line in [
        ["Legende zum Kontoauszug"],
        ["*", "provisorisch verbucht"],
        ["**", "Gutschrift unter Vorbehalt"],
        ["***", "Korrekturbuchung"],
        ["k.A.", "keine Angabe"],
        ["n/a", "nicht anwendbar"],
        ["-", "kein Betrag"],
        ["(1'234.55)", "Belastung, Betrag in Klammern"],
        ["1'234.%s" % ENDASH, "Betrag ohne Rappen"],
    ]:
        legende.append(line)

    # Deterministic docProps so reruns hash identically.
    fixed = datetime.datetime(2026, 1, 1, 0, 0, 0)
    wb.properties.creator = "tdy testdata generator"
    wb.properties.lastModifiedBy = "tdy testdata generator"
    wb.properties.created = fixed
    wb.properties.modified = fixed

    p = out("04_bankauszug.xlsx")
    wb.save(p)
    normalize_zip(p)
    written(p)


FIXED_STAMP = b"2026-01-01T00:00:00Z"


def normalize_zip(path):
    """Rewrite the .xlsx with fixed zip entry timestamps -> byte-stable output.

    openpyxl overwrites wb.properties.modified with the wall clock at save
    time, so docProps/core.xml is pinned back to FIXED_STAMP here as well.
    """
    with zipfile.ZipFile(path) as zin:
        items = [(info.filename, info.external_attr, zin.read(info.filename))
                 for info in zin.infolist()]
    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, attr, data in items:
            if name == "docProps/core.xml":
                data = re.sub(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                              rb"\g<1>" + FIXED_STAMP + rb"\g<2>", data)
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = attr
            info.create_system = 0
            zout.writestr(info, data)
    os.replace(tmp, path)


# ---------------------------------------------------------------------------
# 05: one row per pathology
# ---------------------------------------------------------------------------

def write_grenzfaelle():
    rows = [
        ["Fall", "CHF Apostroph", "CHF Strich", "EUR Schmalraum",
         "Klammer Negativ", "Prozent", "Fussnote", "Fehlwerte"],
        ["A", "1'000", "1'234.%s" % ENDASH, "1%s234,56" % NBSP,
         "(1.234,56)", "12,5 %", "1234*", "n/a"],
        ["B", "-1'234.56", "85.%s" % ENDASH, "12%s000,00" % NBSP,
         "(0,01)", "0,75%", "1235 *", "-"],
        ["C", "CHF -50.00", "0.%s" % ENDASH, "999,99",
         "1.000,00", "-1,25 %", "1236**", "k.A."],
        ["D", "-CHF 50.00", "12'000.%s" % ENDASH, "0,01",
         "(12.345,67)", "100 %", "1237", "#N/A"],
        ["E", "12'345'678.90", "7.%s" % ENDASH, "1%s000%s000,00" % (NBSP, NBSP),
         "0,00", "0 %", "1238 ***", ""],
        ["F", "0.00", "1'000'000.%s" % ENDASH, "-2%s500,50" % NBSP,
         "(1.000.000,00)", "33,33%", "1239*", "1'234.00"],
    ]
    p = out("05_grenzfaelle.csv")
    with open(p, "wb") as fh:
        fh.write(csv_text(rows).encode("utf-8"))
    written(p)


# ---------------------------------------------------------------------------
# 06: irreconcilable locales in one column
# ---------------------------------------------------------------------------

def write_gemischt():
    rows = [
        ["ID", "Betrag", "Quelle"],
        ["1", "1'234.56", "Zuerich"],
        ["2", "1.234,56", "Berlin"],
        ["3", "1234.56", "London"],
        ["4", "1,234.56", "New York"],
        ["5", "1234,56", "Wien"],
        ["6", "12'345.60", "Zuerich"],
    ]
    p = out("06_gemischte_konvention.csv")
    with open(p, "wb") as fh:
        fh.write(csv_text(rows).encode("utf-8"))
    written(p)


def main():
    os.makedirs(OUTDIR, exist_ok=True)
    write_statement_csvs()
    write_dezimalkomma()
    write_xlsx()
    write_grenzfaelle()
    write_gemischt()


if __name__ == "__main__":
    main()
