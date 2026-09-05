#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the fixtures Tasks 1-6 committed ad hoc while fixing defects the
2026-09-03 corpus audit found (see `gap_reports/AUDIT_FINDINGS.md`, gitignored
-- the sweep itself never runs in CI, so anything it *finds* has to become a
committed fixture, exactly as `12_late_surprises.py` did for the previous
audit). This generator's only job is reproducibility: the three files below
already existed as committed bytes before this generator was written, and
regenerating them must not move a single byte, or every sidecar fingerprint
pointing at them goes stale.

Run from the repo root: python3 testdata/gen/15_audit_defects.py

Deterministic; requires openpyxl (and lxml, transitively -- see
gen_fixtures.py's module docstring) for the xlsx fixtures.

--------------------------------------------------------------------------
FIXTURES
--------------------------------------------------------------------------

1. torn_tail_utf8.csv (Task 1, `tests/regression.rs::
   utf8_file_with_a_torn_tail_sample_is_not_mojibake`)

   A CSV that decodes as UTF-8 whole-file but whose last 4096 bytes -- the
   window `read_head_tail` seeks to for the tail half of a sniff sample --
   begin mid multi-byte sequence. The production bug: the sniffer detected
   the *concatenation* of head+tail bytes, so a torn lead byte at the seam
   made a perfectly valid file look like invalid UTF-8 and fall back to
   chardetng, which froze on windows-1252 at confidence 0.80 with no
   warning -- every accented value came back as mojibake.

   Almost every row is ASCII filler, so the file's *shape* proves nothing by
   itself; what matters is that exactly one row -- containing a real
   accented name -- straddles the `len(file) - 4096` boundary at the byte
   that splits its multi-byte character. That row and the final row's
   filler width are what place the tear there; both are asserted below
   rather than trusted, because a generator that merely claims a torn tail
   without checking is the same defect this fixture exists to catch.

2. xl_cell_newline.xlsx (Task 6, `tests/regression.rs::
   a_newline_inside_a_cell_is_not_a_row_boundary`)

   A spreadsheet cell with an embedded newline (`"Manufacture\nOf Beer (In
   Barrels)"`), matching the corpus find `ttb_monthly_stats_2018-12.xlsx`
   cell A10. Investigation found tdy already parses this as ONE row
   correctly -- `render_cell` clones a cell's string verbatim, `\n`
   included, and no code path re-splits row text on it. This fixture pins
   that correct behaviour as a regression test rather than leaving it
   undocumented.

3. xl_money_siblings.xlsx (Task 6, currently unread by any test -- the fix
   is blocked)

   Mirrors `PCA_Report_FY16Q3.xlsx`'s two distinct value-level failure modes
   for the *same* currency number format (`"$"#,##0.00"`), both traced by
   hand against the real file:

     (a) `render_cell`'s `format!("{f}")` on an Excel float is Rust's
         shortest-round-trip Display, which drops a trailing zero
         (100000.10 -> "100000.1") or the decimal point entirely for a whole
         number (100.00 -> "100"). One such value in a column is enough to
         break `guess_type`'s `consistent` (exact scale match) check.
     (b) some of the real file's values already carry IEEE-754 noise baked
         into the stored double *before* tdy ever reads it (a spreadsheet
         formula's accumulated error, e.g. the real file's Total column
         reading back as "42437667.559999995") -- `render_cell` only
         reports it faithfully. That pushes the apparent scale past
         `guess_type`'s `(1..=4)` window, which no column-name override can
         rescue.

   `amount_a` reproduces (a); `amount_b` and `amount_c` reproduce (b) with
   noisy values borrowed verbatim from the real file's Total/Rehabilitation
   columns (perturbed via `v + v*1e-15` so the float64 bit pattern genuinely
   differs from the nearest double to the short decimal literal -- a longer
   decimal literal round-trips cleanly through IEEE-754 no matter how many
   digits you type, and proves nothing). All three keep "amount" in the
   name specifically so a future fix cannot cheat via the existing
   `looks_monetary(name)` heuristic: the point of this defect is that
   money-ness must come from the cell's number format, and none of these
   three currently resolve to Decimal despite every cell here carrying the
   same `"$"#,##0.00"` format.

`xl_money_siblings.xlsx` is read by `tests/regression.rs::
currency_formatted_columns_all_become_decimal` -- `src/xlmoney.rs` reads
`xl/styles.xml` and the sheet XML straight out of the zip, since calamine
0.36 exposes no per-cell number-format through any public API reachable
from tdy's extraction path.

4. xl_money_offset_a.xlsx, xl_money_offset_c.xlsx (`tests/regression.rs::
   money_typing_survives_a_used_range_that_does_not_start_at_column_a`)

   Byte-identical in content -- four columns `label`, `alpha`, `beta`
   (currency-formatted, deliberately named nothing `looks_monetary` would
   catch, with alternating 100.5/200.25 and 300.5/400.25 values so the
   scale is inconsistent and only the cell format can make them decimal),
   `note` -- differing *only* in which sheet column the table starts at:
   `_a` starts at column A, `_c` at column C (two blank leading columns).

   `xlmoney::money_columns` decodes sheet-*absolute* column indices from
   cell references (`<c r="D10">` is column 3, whatever else is on the
   sheet), but a `RawTable`'s column 0 is wherever calamine's used range
   actually starts -- column A only when the sheet's data does too. Before
   `RawTable::col_offset` existed, `_c`'s money columns silently bound
   nothing (or, on an unluckier layout, the wrong column): sheet-absolute
   index 1 (`alpha`, real table column 0) was tested against table column 1
   (`beta`), which is not what a currency format on `alpha` should mean.
   The two files must type identically, which is the point of committing
   them as a pair rather than one file that could pass by accident.
"""

import os
import re
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
OUT = os.path.join(REPO, "testdata")


def note(path, what):
    print(f"wrote {os.path.relpath(path, REPO)} ({os.path.getsize(path)} bytes) - {what}")


_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def repack_deterministic(path):
    """Rewrite an xlsx zip with pinned entry timestamps and patched
    `dcterms:modified` -- same technique as gen_fixtures.py's own
    _repack_deterministic and every other xlsx-writing generator here;
    openpyxl rewrites dcterms:modified at save time whatever wb.properties
    says, so the fixture's bytes (and its blake3 fingerprint) would change
    on every regeneration without this."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        entries = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    entries = [
        (name, _MODIFIED_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", data)
         if name == "docProps/core.xml" else data)
        for name, data in entries
    ]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for name, data in entries:
            info = zipfile.ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            zout.writestr(info, data)
    os.replace(tmp, path)


# ---------------------------------------------------------------------------
# 1. torn_tail_utf8.csv
# ---------------------------------------------------------------------------

TAIL_WINDOW = 4096  # must match sample.rs's WINDOW


def build_torn_tail_utf8():
    header = "code,name,region\n"
    filler_name = "F" * 10
    filler_region = "Region0"

    # Row 1482 (0-indexed): a real accented Andorra parish name, long enough
    # and positioned early enough in the file that its multi-byte characters
    # straddle the `total_size - TAIL_WINDOW` boundary computed below.
    special_index = 1482
    special_name = "Sant Julià de Lòria"  # "Sant Julià de Lòria"
    special_region = "Region3"

    # The final row's filler is widened (16 F's instead of 10) purely to
    # land the file's total size where the tear lines up; this is the one
    # knob a byte-identical regeneration depends on, so it stays a named
    # constant rather than a magic number inline.
    last_index = 1633
    last_name = "F" * 16

    total_rows = 1634

    lines = []
    for i in range(total_rows):
        code = f"AD-{i:04d}"
        if i == special_index:
            name, region = special_name, special_region
        elif i == last_index:
            name, region = last_name, filler_region
        else:
            name, region = filler_name, filler_region
        lines.append(f"{code},{name},{region}")

    text = header + "\n".join(lines) + "\n"
    data = text.encode("utf-8")

    # The whole file must decode as UTF-8 ...
    whole = data.decode("utf-8")
    assert whole == text, "round-trip through utf-8 changed the text"

    # ... but the tail sample -- the last TAIL_WINDOW bytes, exactly what
    # `read_head_tail` seeks to -- must NOT: it has to begin mid multi-byte
    # sequence, or this fixture proves nothing.
    assert len(data) >= 40 * 1024, f"fixture is only {len(data)} bytes, need >= 40 KiB"
    tail = data[-TAIL_WINDOW:]
    try:
        tail.decode("utf-8")
        raise AssertionError(
            "tail sample decoded cleanly -- the tear did not land inside a "
            "multi-byte character; adjust special_index/last_name"
        )
    except UnicodeDecodeError:
        pass  # expected: a torn lead/continuation byte at the tail's start

    out = os.path.join(OUT, "torn_tail_utf8.csv")
    with open(out, "wb") as f:
        f.write(data)
    note(out, "whole-file UTF-8, tail sample begins mid-character")


# ---------------------------------------------------------------------------
# 2 & 3. xl_cell_newline.xlsx, xl_money_siblings.xlsx
# ---------------------------------------------------------------------------


def build_cell_newline():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Category", "Jan", "Feb", "Mar"])
    ws.append(["Widgets", 10, 20, 30])
    ws.append(["Manufacture\nOf Beer (In Barrels)", 40, 50, 60])
    ws.append(["Gadgets", 70, 80, 90])

    out = os.path.join(OUT, "xl_cell_newline.xlsx")
    _save_deterministic(wb, out)
    note(out, "3 data rows, one cell with an embedded newline")


def build_money_siblings():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    fmt = '"$"#,##0.00'

    # A typed decimal literal round-trips cleanly through IEEE-754 no matter
    # how many (correct) digits you type -- Rust's shortest-round-trip
    # Display finds the same short string either way. Real noise like the
    # source file's needs a genuine floating-point *operation* to perturb
    # the bit pattern away from the nearest-double-to-a-short-decimal.
    def noisy(v):
        return v + v * 1e-15

    ws.append(["agency_name", "added", "amount_a", "amount_b", "amount_c"])
    rows = [
        ["Agency One", 255871181.11, noisy(11147705.78), noisy(42437667.56), noisy(255871181.11)],
        ["Agency Two", 100000.10, noisy(10977014.90), noisy(10343512.58), noisy(7340146.63)],
        ["Agency Three", 3.33, 38180001.97, 9.99, 12.12],
    ]
    for r in rows:
        ws.append(r)
    for row in range(2, 2 + len(rows)):
        for col in range(2, 6):  # added, amount_a, amount_b, amount_c
            ws.cell(row=row, column=col).number_format = fmt

    out = os.path.join(OUT, "xl_money_siblings.xlsx")
    _save_deterministic(wb, out)
    note(out, "4 currency-formatted sibling columns, lossy float64 values")


def _build_money_offset(first_col, filename, what):
    """One workbook for the money/column-offset pair: four columns starting
    at `first_col` (1 = A, 3 = C), `alpha`/`beta` currency-formatted with
    names that `looks_monetary` would not catch and values whose scale is
    inconsistent (100.5 vs 200.25) so only the cell format can make them
    decimal. Column layout is otherwise irrelevant to the point of the
    fixture -- what matters is that `_a` and `_c` disagree only on
    `first_col`."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    fmt = '"$"#,##0.00'
    for j, h in enumerate(["label", "alpha", "beta", "note"]):
        ws.cell(row=1, column=first_col + j, value=h)
    for i in range(30):
        ws.cell(row=2 + i, column=first_col, value=f"row {i}")
        a = ws.cell(row=2 + i, column=first_col + 1, value=100.5 if i % 2 else 200.25)
        b = ws.cell(row=2 + i, column=first_col + 2, value=300.5 if i % 2 else 400.25)
        a.number_format = fmt
        b.number_format = fmt
        ws.cell(row=2 + i, column=first_col + 3, value=f"n{i}")

    out = os.path.join(OUT, filename)
    _save_deterministic(wb, out)
    note(out, what)


def build_money_offset_pair():
    _build_money_offset(1, "xl_money_offset_a.xlsx", "money columns starting at sheet column A")
    _build_money_offset(
        3, "xl_money_offset_c.xlsx", "byte-for-byte the same table, starting at sheet column C"
    )


def build_ragged_header():
    """A wide, ragged sheet whose real header is row 1 and whose FIRST DATA
    ROW is sparse enough to disqualify it.

    `sniff_excel_sheet` picks the header as "the first row with >= 60% of the
    grid width populated that is followed by a row with >= 50%". On a sheet
    where institutions report different numbers of sports -- rows ranging
    from a handful of cells to the full grid -- row 1 is a perfect header but
    row 2 falls under the 50% bar, so the scan walks *into the data* and
    stops at the first pair of rows that happen to be dense enough. It then
    finds the row it landed on is not a header, declines to promote it, and
    keeps the skip anyway: N rows deleted for a reason that was withdrawn.

    Found on the EADA `Schools.xlsx` family (2022-03-29 tidytuesday), where
    four of five workbooks lost rows and the worst lost 216 -- seven real
    universities. Reproduced here at 20 columns so the arithmetic is legible:

        row  1 : 20 populated  (>= 12 = 60%)  but row 2 has 9 (< 10 = 50%)
        rows 2-9: 8-9 populated                       -> never qualify
        row 10 : 13 populated  (>= 12)  and row 11 has 11 (>= 10)  -> PICKED

    so `header_idx` = 9, `skip_rows{head=9}` deletes the header and the eight
    data rows above it, and the columns come back `col_1, col_2, ...`.

    The fixture's contract: 29 data rows, all of them readable, under the
    header's own names.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "EADA_schools"

    W = 20
    header = ["unitid", "institution_name"] + [f"sport_{i}_participants" for i in range(1, W - 1)]
    assert len(header) == W, len(header)
    ws.append(header)

    def row(n, metrics):
        """One institution reporting `metrics` sports; the rest of the grid
        is genuinely empty, which is what makes the sheet ragged."""
        cells = [100000 + n, f"University {n}"] + [10 * n + k for k in range(metrics)]
        return cells + [None] * (W - len(cells))

    # Row 2: 9 populated -- two short of the 50% bar, which is the whole
    # defect. Rows 3-9: 8 populated. Row 10: 13, row 11: 11 -> the pair the
    # scan stops at.
    plan = [7] + [6] * 7 + [11, 9] + [6] * 19
    assert len(plan) == 29, len(plan)
    for n, metrics in enumerate(plan, start=1):
        ws.append(row(n, metrics))

    populated = [sum(1 for c in r if c is not None and c != "") for r in ws.iter_rows(values_only=True)]
    assert populated[0] == W, populated[0]
    assert populated[1] == 9, populated[1]          # first data row, under 50%
    assert populated[9] == 13, populated[9]         # >= 60%
    assert populated[10] == 11, populated[10]       # >= 50%
    assert len(populated) == 30, len(populated)

    out = os.path.join(OUT, "xl_ragged_header.xlsx")
    _save_deterministic(wb, out)
    note(out, "wide ragged sheet; real header row 1, sparse first data row")


def _save_deterministic(wb, out):
    """Pin document properties and repack the zip with fixed entry stamps,
    exactly as gen_fixtures.py's own _repack_deterministic does -- openpyxl
    rewrites dcterms:modified at save time whatever wb.properties says, so
    without this every regeneration would stale the blake3 fingerprint in
    any committed sidecar pointing at the file."""
    from datetime import datetime

    wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
    wb.save(out)
    repack_deterministic(out)


def main():
    os.makedirs(OUT, exist_ok=True)
    build_torn_tail_utf8()
    build_cell_newline()
    build_money_siblings()
    build_money_offset_pair()
    build_ragged_header()


if __name__ == "__main__":
    main()
