#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the `excel_nightmares_*` fixture family for tdy.

Run from the repo root:  python3 testdata/gen/02_excel_nightmares.py

Deterministic and idempotent: randomness is seeded, document properties are
pinned and each workbook is repacked as a byte-stable zip, so re-running the
generator produces identical bytes and does not silently stale a committed
sidecar's blake3 fingerprint.

Everything here is a *structural* nightmare rather than a corrupt file: every
workbook opens fine in Excel and in calamine, and is wrong only if the
parsing spec is wrong. Each fixture's "correct parse" below was executed
against the real engine (target/debug/tdy with a hand-written sidecar), and
each failure mode was executed too — the numbers are measured, not predicted.

Ground truth this family depends on (calamine 0.26 through
`src/sample.rs::render_cell`, i.e. the exact strings the executor sees):

  * a real Excel date cell (serial + a date `number_format`) renders as
    "YYYY-MM-DD HH:MM:SS", NOT "YYYY-MM-DD": dtype date with format
    "%Y-%m-%d" errors, while "%Y-%m-%d %H:%M:%S" parses (chrono discards the
    time), as does a `strip` of " 00:00:00$";
  * a formula cell written by openpyxl carries <f> with an EMPTY <v>
    (openpyxl computes nothing), so calamine yields Data::Empty and the
    column arrives ALL NULL — never the formula text, never the value;
  * an error cell renders as the literal string "#ERR:Div0"
    (Data::Error(Div0)), so a numeric column holding one needs
    na_values = ["#ERR:Div0"];
  * a whitespace-only cell extends calamine's used range in both directions;
    an empty-string cell and a styled-but-valueless cell do NOT;
  * leading empty rows/columns are trimmed from the used range (a sheet
    whose first cell is C5 starts at C5); trailing ones extend only to the
    last genuinely non-empty cell;
  * `promote_header` trims each header cell, so a header written as
    "Betrag CHF " is addressed as "Betrag CHF"; an embedded newline survives
    ("Konto\n(Soll)").

--------------------------------------------------------------------------
FIXTURES  (all in testdata/, all named excel_nightmares_*)
--------------------------------------------------------------------------

1. excel_nightmares_cover_sheet.xlsx
   Stresses sheet selection. Sheets are ["Deckblatt", "Hinweise",
   "Bewegungen"]: sheet 0 is a prose cover page, sheet 1 is a two-column
   glossary that *looks* tabular (the decoy), the data is on sheet 2.
   `src/sample.rs` renders only the first two sheets, so the LLM tier never
   sees the data at all — only the sheet-name list points at it. Header
   cells carry a trailing space and an embedded newline.
   Correct parse: extraction sheet_name = "Bewegungen" (or sheet_index = 2)
   + promote_header rows = 1 → 24 rows x 5 columns
   [buchungs_nr Int64, datum Date, konto_soll Utf8 (source "Konto\n(Soll)"),
   betrag_chf Decimal(12,2) (source "Betrag CHF"), waehrung Utf8];
   row 0 = (1001, 2025-01-02, "4400 Erloese"/"4400 Erlöse", 1234.50, "CHF");
   sum(betrag_chf) = 56469.00; 3 rows have waehrung = 'EUR';
   count(distinct konto_soll) = 4.
   Measured tier-1 behaviour: picks sheet 0 and reports confidence 0.90 for
   a single Utf8 column of cover-page prose — high confidence, wrong sheet.

2. excel_nightmares_merged_header.xlsx
   Stresses a three-row merged header, vertically merged category cells,
   interleaved group subtotals, blank separator rows, a grand total, and
   transform ORDER.
   Correct parse: skip_rows{head:2, tail:1} → promote_header{rows:3} (fills
   right, then joins) yielding exactly
     ["Region", "Produkt", "2025 Q1 Jan", "2025 Q1 Feb", "2025 Q1 Mär",
      "2025 Q2 Apr", "2025 Q2 Mai", "2025 Q2 Jun"]
   → drop_rows_matching "^\t*$" (the blank separator rows, whole-row match)
   → drop_rows_matching "^Zwischensumme" on column Produkt (the subtotal
   rows leave Region EMPTY on purpose, so dropping on Region cannot work)
   → fill_down["Region"] → unpivot the six month columns.
   Result: 54 rows x 4 columns [region Utf8, produkt Utf8, monat Date,
   stueck Int64]; sum(stueck) = 95877; per region Ost 24561 / West 31959 /
   Süd 39357, 18 rows each; (Ost, Widget, 2025-01-01) = 1200.
   `monat` needs parse.replace [Mär→Mar, Mai→May] and parse.strip " Q[0-9]"
   before dtype date "%Y %b" — exercising the trim→replace→na→strip order
   (Jan/Feb/Apr/Jun happen to be spelled the same in German).
   Measured failure modes: leaving the subtotal rows in gives exactly 72
   rows and 191754 = 2 x 95877; running fill_down BEFORE dropping the blank
   rows makes the blanks survive the "^\t*$" test (Region is filled by then)
   and errors with "row 19: null in non-nullable column".
   Measured tier-1 behaviour: confidence 0.60 and a spec its own engine
   cannot execute — guess_columns emits ColumnSpec{source: ""} for the blank
   header cells while promote_header has already filled them right and
   deduped, so execution dies with `no column named ""`.

3. excel_nightmares_mixed_dates.xlsx
   Stresses date representations that cannot be reconciled, plus an ID
   column of zero-padded digit strings. 12 rows x 5 columns.
   Correct parse: `vertrag_nr` MUST stay Utf8 → row 0 = "00123" (Int64 turns
   it into 123); `start` is Date "%Y-%m-%d %H:%M:%S" → row 0 = 2025-01-02;
   `tage` Int64 with sum = 2340.
   `ende` mixes 4 real Excel dates, 4 text dates ("31.12.2025", …), 2
   "offen" and 2 blanks: this grammar has no per-value alternative format,
   so the only correct outcomes are Utf8 (8 non-null with
   na_values=["offen"]) or a hard error — typing it as a date fails loudly
   at "row 5: cannot parse \"31.12.2025\"", which is the invariant-6
   behaviour to preserve.
   `aenderung` is the silent one: all text dd/mm/yyyy with every component
   <= 12, so BOTH readings parse the whole column without error — measured:
   "%d/%m/%Y" → 2025-04-03, "%m/%d/%Y" → 2025-03-04 for row 0. The sniffer
   picks "%m/%d/%Y" purely by list order and reports confidence 0.90, so a
   German workbook (umlauts, DD.MM.YYYY number formats, "Vertrag") is read
   as US dates with nothing flagged.
   Measured tier-1 behaviour also types vertrag_nr as Int64 ("00123" → 123)
   and names the column "nderung", because `sanitize` maps only lowercase
   umlauts and drops the leading "Ä".

4. excel_nightmares_formulas.xlsx
   Stresses formula cells, a real error cell, and numbers stored as text
   with a German decimal comma. 8 rows x 6 columns.
   Correct parse: `summe_formel` is ALL NULL (0 of 8 non-null) — openpyxl
   caches no value, so the engine must not invent one; with nullable=false
   it must error ("row 1: null in non-nullable column"). `marge` holds a
   genuine Excel error cell in row index 2, rendered "#ERR:Div0", so Float64
   needs na_values=["#ERR:Div0"] → 7 of 8 non-null. `betrag_text` holds
   "1,5", "2,75", "10,25", "1234,00", "0,99", "45,60", "7,05", "310,40" as
   text: the correct spec is decimal_separator = ',' → row 0 = 1.50 and
   sum = 1612.54.
   Measured tier-1 behaviour (the corruption bomb): guess_type tries
   separator conventions "first one that parses" and picks
   thousands_separator = ',', producing 15.0, 275.0, 1025.0, 123400.0, 99.0,
   4560.0, 705.0, 31040.0 — sum 161119.0, row 0 ten times too large, no
   error anywhere, reported confidence 0.90. src/numfmt.rs was written to
   prevent exactly this and is not declared in src/lib.rs, so it never runs.

5. excel_nightmares_empty_sheet.xlsx
   Stresses degenerate sheets. Sheets are ["Leer", "Fast leer", "Daten"]:
   sheet 0 has no cells at all (<sheetData></sheetData>), sheet 1 holds a
   single whitespace cell at D9 and nothing else, sheet 2 holds a clean
   4-row table.
   Correct parse: on the default sheet the only correct outcome is a clear
   error — measured: `tdy sniff` fails with "sheet appears to be empty" and
   writes no sidecar. With sheet_name = "Daten": 4 rows x 3 columns
   [code Utf8, menge Int64, farbe Utf8]; row 0 = ("A-1", 10, "rot");
   sum(menge) = 100; 2 rows have farbe = 'rot'.
   The whitespace-only sheet is its own trap: its used range is the single
   cell D9, so the sniffer promotes " " to a header, keeps zero body rows,
   and emits a spec that cannot execute (`no column named " "`).

6. excel_nightmares_phantom_range.xlsx
   Stresses a used range inflated by one stray whitespace cell at H400 while
   the real data lives in A1:D21 (dimension is A1:H400).
   Measured naive read (no range): 399 body rows and 8 columns, of which 379
   rows and 4 columns are entirely empty — sum(betrag) is still 4570.00, so
   only the row count betrays the problem.
   Correct parse: range = "A1:D21" (or dropping all-empty rows) → 20 rows x
   4 columns [beleg Utf8, datum Date, betrag Decimal(12,2), bemerkung Utf8];
   row 0 = ("B-0001", 2025-02-03, 100.25, "Skonto"); sum(betrag) = 4570.00;
   7 of 20 rows have a non-empty bemerkung.
   The four phantom header cells are blank, which again makes tier-1 emit
   ColumnSpec{source: ""} against a filled-right, deduped header — a spec
   the engine refuses (invariant 1).

7. excel_nightmares_3000_rows.xlsx
   Stresses size and a trailing grand total: 3000 body rows plus a "Total"
   footer, 92 KB on disk.
   Correct parse: skip_rows{tail:1} + promote_header → exactly 3000 rows x 5
   columns [id Int64, datum Date, region Utf8, kanal Utf8, betrag
   Decimal(12,2)]; row 0 = (1, 2024-01-01, "Ost", "Filiale", 1655.22);
   sum(betrag) = 7627533.57; min(datum) = 2024-01-01, max = 2024-12-31;
   4 distinct regions, 3 distinct channels.
   The footer carries the true total, so forgetting tail:1 doubles the
   answer: measured 3001 rows and sum 15255067.14 = 2 x 7627533.57 when `id`
   is Utf8. With `id` typed Int64 the same mistake errors instead
   ("row 3001: cannot parse \"Total\""), which is the invariant-6 outcome
   worth keeping.
"""

from __future__ import annotations

import datetime as dt
import os
import random
import re
import zipfile
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

SEED = 20260828
PREFIX = "excel_nightmares_"
OUT_DIR = Path(__file__).resolve().parents[1]          # <repo>/testdata
FIXED_TS = dt.datetime(2026, 1, 1, 0, 0, 0)
DATE_FMT = "DD.MM.YYYY"
TEXT_FMT = "@"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def new_workbook(first_sheet: str) -> Workbook:
    wb = Workbook()
    wb.active.title = first_sheet
    wb.properties.created = FIXED_TS
    wb.properties.modified = FIXED_TS
    wb.properties.creator = "tdy fixture generator"
    wb.properties.lastModifiedBy = "tdy fixture generator"
    return wb


def date_cell(ws, row: int, col: int, value: dt.date) -> None:
    """A *real* Excel date: a serial number plus a date number_format."""
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = DATE_FMT


def text_cell(ws, row: int, col: int, value: str) -> None:
    """A cell explicitly formatted as text (numbers stored as text)."""
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = TEXT_FMT


MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def repack_deterministic(path: Path) -> None:
    """Rewrite the xlsx zip with pinned entry timestamps and permissions.

    openpyxl stamps each zip entry with the wall clock and overwrites
    `dcterms:modified` at save time, which would make the generator
    non-idempotent at the byte level and silently stale every committed
    sidecar's blake3 fingerprint on each regeneration.
    """
    tmp = path.with_suffix(".xlsx.tmp")
    with zipfile.ZipFile(path) as zin:
        entries = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    entries = [
        (name, MODIFIED_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", data)
         if name == "docProps/core.xml" else data)
        for name, data in entries
    ]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for name, data in entries:
            info = zipfile.ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            zout.writestr(info, data)
    os.replace(tmp, path)


def save(wb: Workbook, name: str, note: str) -> None:
    path = OUT_DIR / (PREFIX + name)
    wb.save(path)
    repack_deterministic(path)
    print(f"wrote {path.relative_to(OUT_DIR.parent)} ({path.stat().st_size} bytes) - {note}")


# ---------------------------------------------------------------------------
# 1. cover page on sheet 0, decoy glossary on sheet 1, data on sheet 2
# ---------------------------------------------------------------------------

KONTEN = ["4400 Erlöse", "6500 Verwaltung", "3200 Material", "6000 Personal"]


def build_cover_sheet() -> None:
    wb = new_workbook("Deckblatt")

    cover = wb["Deckblatt"]
    cover["B2"] = "Muster AG"
    cover["B3"] = "Buchungsjournal 2025"
    cover["B5"] = "Vertraulich — nur für den internen Gebrauch"
    cover["B6"] = "Erstellt am 07.01.2026 durch Abteilung Controlling"
    cover["B8"] = ("Dieses Dokument enthält Bewegungsdaten des Geschäftsjahres "
                   "2025. Die Auswertung befindet sich im Register 'Bewegungen'.")
    cover["B10"] = "Kontakt: controlling@muster.example"
    cover.merge_cells("B3:E3")
    cover.merge_cells("B8:E8")

    # A decoy: two columns, looks perfectly tabular to a naive sheet picker.
    legend = wb.create_sheet("Hinweise")
    legend.append(["Feld", "Bedeutung"])
    for field, meaning in [
        ("Buchungs-Nr", "fortlaufende Nummer"),
        ("Datum", "Buchungsdatum"),
        ("Konto (Soll)", "Kontonummer und Bezeichnung"),
        ("Betrag CHF", "Bruttobetrag in Schweizer Franken"),
        ("Währung", "Originalwährung des Belegs"),
        ("Storno", "wird in diesem Export nicht ausgewiesen"),
    ]:
        legend.append([field, meaning])

    data = wb.create_sheet("Bewegungen")
    # Header horrors: a trailing space, and a newline inside a header cell.
    data.append(["Buchungs-Nr", "Datum", "Konto\n(Soll)", "Betrag CHF ", "Währung"])
    total = Decimal("0.00")
    for i in range(24):
        row = i + 2
        data.cell(row=row, column=1, value=1001 + i)
        date_cell(data, row, 2, dt.date(2025, 1, 2) + dt.timedelta(days=13 * i))
        data.cell(row=row, column=3, value=KONTEN[i % len(KONTEN)])
        betrag = (Decimal("1234.50") + Decimal("97.25") * i).quantize(Decimal("0.01"))
        total += betrag
        data.cell(row=row, column=4, value=float(betrag))
        data.cell(row=row, column=5, value="EUR" if i % 7 == 3 else "CHF")
    assert total == Decimal("56469.00"), total

    save(wb, "cover_sheet.xlsx",
         "cover on sheet 0, decoy glossary on sheet 1, 24 data rows on sheet 2")


# ---------------------------------------------------------------------------
# 2. three-row merged header + merged categories + subtotals + grand total
# ---------------------------------------------------------------------------

REGIONS = ["Ost", "West", "Süd"]
PRODUCTS = ["Widget", "Gadget", "Doohickey"]
MONTHS = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun"]


def units(region_idx: int, product_idx: int, month_idx: int) -> int:
    return 1200 + 137 * (3 * region_idx + product_idx) + 11 * month_idx


def build_merged_header() -> None:
    wb = new_workbook("Absatz")
    ws = wb["Absatz"]

    ws["A1"] = "Muster AG — Absatzstatistik 2025"
    ws.merge_cells("A1:H1")
    ws["A2"] = "Alle Angaben in Stück; erstellt 2026-01-07"
    ws.merge_cells("A2:H2")

    # Three-row header. Row 3: year over the six month columns. Row 4:
    # quarters over three months each. Row 5: month names.
    ws["A3"] = "Region"
    ws["B3"] = "Produkt"
    ws["C3"] = "2025"
    ws.merge_cells("A3:A5")
    ws.merge_cells("B3:B5")
    ws.merge_cells("C3:H3")
    ws["C4"] = "Q1"
    ws.merge_cells("C4:E4")
    ws["F4"] = "Q2"
    ws.merge_cells("F4:H4")
    for m, month in enumerate(MONTHS):
        ws.cell(row=5, column=3 + m, value=month)

    grand = [0] * len(MONTHS)
    row = 6
    for r, region in enumerate(REGIONS):
        block_start = row
        subtotal = [0] * len(MONTHS)
        for p, product in enumerate(PRODUCTS):
            ws.cell(row=row, column=1, value=region if p == 0 else None)
            ws.cell(row=row, column=2, value=product)
            for m in range(len(MONTHS)):
                v = units(r, p, m)
                subtotal[m] += v
                grand[m] += v
                ws.cell(row=row, column=3 + m, value=v)
            row += 1
        # Vertically merged category cell: the value lives only in the top row.
        ws.merge_cells(start_row=block_start, start_column=1,
                       end_row=row - 1, end_column=1)
        # Subtotal row: column A stays EMPTY, so it cannot be dropped on Region.
        ws.cell(row=row, column=2, value=f"Zwischensumme {region}")
        for m in range(len(MONTHS)):
            ws.cell(row=row, column=3 + m, value=subtotal[m])
        row += 1
        if r < len(REGIONS) - 1:
            row += 1  # a completely blank separator row

    ws.cell(row=row, column=1, value="Total")
    for m in range(len(MONTHS)):
        ws.cell(row=row, column=3 + m, value=grand[m])

    assert sum(grand) == 95877, sum(grand)
    save(wb, "merged_header.xlsx",
         "3-row merged header, merged regions, 3 subtotals + blank rows + grand total")


# ---------------------------------------------------------------------------
# 3. real Excel dates vs. text dates vs. ambiguous dd/mm text dates
# ---------------------------------------------------------------------------

TEXT_ENDS = ["31.12.2025", "15.01.2026", "28.02.2026", "31.03.2026"]
# Every component <= 12, so %d/%m/%Y and %m/%d/%Y both parse the whole column.
AMBIGUOUS = [
    "03/04/2025", "05/06/2025", "11/12/2025", "01/02/2025",
    "07/08/2025", "09/10/2025", "02/03/2025", "04/05/2025",
    "06/07/2025", "08/09/2025", "10/11/2025", "12/01/2026",
]


def build_mixed_dates() -> None:
    wb = new_workbook("Termine")
    ws = wb["Termine"]
    ws.append(["Vertrag-Nr", "Start", "Ende", "Änderung", "Tage"])

    for i in range(12):
        row = i + 2
        text_cell(ws, row, 1, f"{123 + i:05d}")            # "00123": Int64 loses it
        date_cell(ws, row, 2, dt.date(2025, 1, 2) + dt.timedelta(days=30 * i))
        if i < 4:                                          # real Excel dates
            date_cell(ws, row, 3, dt.date(2025, 6, 30) + dt.timedelta(days=15 * i))
        elif i < 8:                                        # text dates
            text_cell(ws, row, 3, TEXT_ENDS[i - 4])
        elif i < 10:                                       # a word, not a date
            ws.cell(row=row, column=3, value="offen")
        # i >= 10: left blank on purpose
        text_cell(ws, row, 4, AMBIGUOUS[i])
        ws.cell(row=row, column=5, value=30 * (i + 1))

    save(wb, "mixed_dates.xlsx",
         "real dates + text dates + 'offen' + blanks, and an unresolvable dd/mm column")


# ---------------------------------------------------------------------------
# 4. formulas, an error cell, and numbers stored as text (German comma)
# ---------------------------------------------------------------------------

MENGEN = [3, 7, 0, 12, 5, 9, 4, 6]
PREISE = [12.5, 8.0, 99.9, 4.25, 33.0, 1.75, 250.0, 19.9]
MARGEN = [0.21, 0.18, None, 0.33, 0.27, 0.12, 0.4, 0.09]   # None -> #DIV/0!
BETRAG_TEXT = ["1,5", "2,75", "10,25", "1234,00", "0,99", "45,60", "7,05", "310,40"]


def build_formulas() -> None:
    wb = new_workbook("Kalkulation")
    ws = wb["Kalkulation"]
    ws.append(["Position", "Menge", "Preis CHF", "Summe (Formel)", "Marge", "Betrag Text"])

    for i in range(8):
        row = i + 2
        ws.cell(row=row, column=1, value=f"P-{i + 1:03d}")
        ws.cell(row=row, column=2, value=MENGEN[i])
        ws.cell(row=row, column=3, value=PREISE[i])
        # openpyxl stores the formula with an EMPTY cached <v>: reads as null.
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
        if MARGEN[i] is None:
            # A genuine Excel error cell (openpyxl writes t="e").
            ws.cell(row=row, column=5, value="#DIV/0!")
        else:
            ws.cell(row=row, column=5, value=MARGEN[i])
        text_cell(ws, row, 6, BETRAG_TEXT[i])

    correct = sum(Decimal(v.replace(",", ".")) for v in BETRAG_TEXT)
    assert correct == Decimal("1612.54"), correct
    corrupt = sum(float(v.replace(",", "")) for v in BETRAG_TEXT)
    assert corrupt == 161119.0, corrupt

    save(wb, "formulas.xlsx",
         "formula column (reads all-null), a #DIV/0! cell, decimal-comma text numbers")


# ---------------------------------------------------------------------------
# 5. a completely empty sheet, a whitespace-only sheet, then real data
# ---------------------------------------------------------------------------

def build_empty_sheet() -> None:
    wb = new_workbook("Leer")           # sheet 0: not a single cell
    almost = wb.create_sheet("Fast leer")
    almost["D9"] = " "                  # one whitespace cell, nothing else

    data = wb.create_sheet("Daten")
    data.append(["Code", "Menge", "Farbe"])
    for i, (code, menge, farbe) in enumerate(
        [("A-1", 10, "rot"), ("A-2", 20, "blau"), ("A-3", 30, "grün"), ("A-4", 40, "rot")]
    ):
        data.append([code, menge, farbe])

    save(wb, "empty_sheet.xlsx",
         "sheet 0 has no cells, sheet 1 has one blank cell, sheet 2 has 4 rows")


# ---------------------------------------------------------------------------
# 6. a used range inflated by one stray whitespace cell
# ---------------------------------------------------------------------------

BEMERKUNGEN = ["Skonto", "", "Teillieferung", "", "", "Retoure", "", "", "Skonto", "",
               "", "Nachbelastung", "", "", "", "Gutschrift", "", "", "", "Storno"]


def build_phantom_range() -> None:
    wb = new_workbook("Export")
    ws = wb["Export"]
    ws.append(["Beleg", "Datum", "Betrag", "Bemerkung"])

    total = Decimal("0.00")
    for i in range(20):
        row = i + 2
        ws.cell(row=row, column=1, value=f"B-{i + 1:04d}")
        date_cell(ws, row, 2, dt.date(2025, 2, 3) + dt.timedelta(days=7 * i))
        betrag = (Decimal("100.25") + Decimal("13.50") * i).quantize(Decimal("0.01"))
        total += betrag
        ws.cell(row=row, column=3, value=float(betrag))
        if BEMERKUNGEN[i]:
            ws.cell(row=row, column=4, value=BEMERKUNGEN[i])

    # The whole trick: one whitespace cell far away. An empty string or a
    # style alone would NOT extend calamine's used range; a space does.
    ws["H400"] = " "

    assert total == Decimal("4570.00"), total
    save(wb, "phantom_range.xlsx",
         "real data in A1:D21, one stray ' ' at H400 inflates the range to A1:H400")


# ---------------------------------------------------------------------------
# 7. 3000 rows plus a grand total footer
# ---------------------------------------------------------------------------

BIG_REGIONS = ["Ost", "West", "Süd", "Nord"]
KANAELE = ["Web", "Filiale", "Partner"]


def build_big_sheet() -> None:
    rng = random.Random(SEED)
    wb = new_workbook("Rohdaten")
    ws = wb["Rohdaten"]
    ws.append(["ID", "Datum", "Region", "Kanal", "Betrag"])

    total = Decimal("0.00")
    for i in range(3000):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        date_cell(ws, row, 2, dt.date(2024, 1, 1) + dt.timedelta(days=i % 366))
        ws.cell(row=row, column=3, value=BIG_REGIONS[i % len(BIG_REGIONS)])
        ws.cell(row=row, column=4, value=rng.choice(KANAELE))
        betrag = Decimal(str(round(rng.uniform(10.0, 5000.0), 2)))
        total += betrag
        ws.cell(row=row, column=5, value=float(betrag))

    footer = 3002
    ws.cell(row=footer, column=1, value="Total")
    ws.cell(row=footer, column=5, value=float(total))

    save(wb, "3000_rows.xlsx",
         f"3000 data rows + a 'Total' footer; sum(Betrag) = {total}")


# ---------------------------------------------------------------------------

def main() -> None:
    random.seed(SEED)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    build_cover_sheet()
    build_merged_header()
    build_mixed_dates()
    build_formulas()
    build_empty_sheet()
    build_phantom_range()
    build_big_sheet()


if __name__ == "__main__":
    main()
