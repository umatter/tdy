#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the `drifting_exports/` fixture family for tdy (job key: drifting-exports).

Run from the repo root:  python3 testdata/gen/11_drifting_exports.py

Deterministic and idempotent; openpyxl for the two workbooks, stdlib for the
CSVs. Everything lands in testdata/drifting_exports/.

WHY THIS FAMILY EXISTS
--------------------------------------------------------------------------
Every other fixture family stresses *one file*. This one stresses a **pile**:
twelve monthly exports of the same logical dataset, drifting the way real
monthly exports drift, plus the target that declares what they should all
become. It is the scenario `docs/design/2026-08-30-target-schema.md` is written
against, and it is the corpus `tdy fit` will be judged on.

The point is not that each file is hard. Most of them are easy. The point is
that they disagree with *each other* — and that a tool which quietly papers over
one month's disagreement produces a number that is short by a twelfth, is
well-typed, raises no error, and survives review. That is the failure this
whole layer exists to prevent, and it cannot be demonstrated with one file.

THE DECLARED TARGET  (drifting_exports/sales.tdy.sql)
--------------------------------------------------------------------------
    month      DATE          NOT NULL
    region     TEXT          NOT NULL
    amount_chf DECIMAL(14,2) NOT NULL

--------------------------------------------------------------------------
FIXTURES
--------------------------------------------------------------------------
Amounts are exact and stated per file below; the whole point of this family is
that the arithmetic is checkable.

1. 2025-01.csv … 2025-06.csv  — the ordinary six.
   `Datum;Region;Betrag`, semicolon-delimited, **windows-1252**, Swiss
   apostrophe grouping (`1'234.50`), day-first dates (`31.01.2025`).
   These must fit mechanically: every declared column matches a header cell by
   normalised name, and the types check.

2. 2025-07.csv  — THE UNIT TRAP.
   Identical layout, except the amount column is named `Betrag Rp.` and holds
   **integer Rappen**: 123450, not 1234.50. Nothing about it is malformed. It
   parses. Bound naively to `amount_chf` it is wrong by a factor of 100, and
   the error is invisible in any single row.
   Correct behaviour today (no `decimal_shift` operator yet): `Betrag Rp.`
   does not normalise to `Betrag`, so nothing binds and the file is a gap. The
   refusal is the correct answer, not a limitation.

3. 2025-08.csv  — THE AMBIGUITY TRAP.
   Two columns named literally `Betrag` — net and gross, in that order. A
   binder that takes the first match is right half the time and silent about
   it. Correct behaviour: refuse, and say which two columns collided.

4. 2025-09.xlsx  — the structural one.
   Sheet "Umsatz": a title row, a merged band `Umsatz 2025` over C:E, then the
   real header `Datum | Region | Betrag CHF`. Needs skip_rows + promote_header
   before any column name exists to match against — which is why the planner
   enumerates frames before it binds.

5. 2025-10.xlsx  — the English locale.
   `Date, Region, Amount, Discount`: different header names, ISO dates
   (`2025-10-31`), plain decimal point, and an extra column the target does not
   declare. Exercises alias matching, and confirms an undeclared column is
   dropped rather than an error.

6. 2025-11.csv  — THE PARTIAL EXPORT.
   `Datum;Betrag` only — no region column at all. There is no plan that reaches
   the target. Correct behaviour: a hard error naming `region`, not a load with
   a null-filled column. A dataset that is quietly short one column is exactly
   the aggregate-laundering failure the design refuses.

7. 2025-12.csv  — the harmless extra.
   Ordinary layout plus a `Kundennummer` column. Must fit, with the extra
   column dropped by the projection.

GROUND TRUTH  (printed by this script, and asserted in tests)
--------------------------------------------------------------------------
Each month has 4 rows, one per region in the order Ost, West, Nord, Sued.
The amount for month m (1-based), row r (0-based) is:

    1000.00 + 100*m + 10*r          e.g. 2025-03 row 2 -> 1320.00

so every month sums to 4060 + 400*m, and the **nine** fittable months
(1-6, 9, 10, 12) sum to **57_340.00**. July is in Rappen, August is ambiguous
and November is short a column, so none of those three may join.
"""

import os
import re
import zipfile
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
OUT = os.path.join(REPO, "testdata", "drifting_exports")

ZIP_EPOCH = (2026, 1, 1, 0, 0, 0)
REGIONS = ["Ost", "West", "Nord", "Sued"]
# Last day of each month in 2025, so the dates are real and month-distinct.
LAST_DAY = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


def amount(month, row):
    """Exact, stated in the docstring, and checkable by hand."""
    return 1000.00 + 100 * month + 10 * row


def rows(month):
    return [(REGIONS[r], amount(month, r)) for r in range(4)]


def swiss(v):
    """1234.5 -> 1'234.50"""
    whole, frac = f"{v:.2f}".split(".")
    grouped = ""
    while len(whole) > 3:
        grouped = "'" + whole[-3:] + grouped
        whole = whole[:-3]
    return f"{whole}{grouped}.{frac}"


def note(path, what):
    print(f"wrote {os.path.relpath(path, REPO)} ({os.path.getsize(path)} bytes) - {what}")


def write_cp1252(name, text, what):
    path = os.path.join(OUT, name)
    with open(path, "wb") as f:
        f.write(text.encode("windows-1252"))
    note(path, what)


def build_ordinary_months():
    """The six that must fit with no ceremony."""
    for m in range(1, 7):
        lines = ["Datum;Region;Betrag"]
        for region, amt in rows(m):
            lines.append(f"{LAST_DAY[m - 1]:02d}.{m:02d}.2025;{region};{swiss(amt)}")
        write_cp1252(
            f"2025-{m:02d}.csv",
            "\n".join(lines) + "\n",
            f"ordinary; sum {sum(a for _, a in rows(m)):.2f}",
        )


def build_rappen():
    """Integer Rappen behind a column name that does not normalise to Betrag."""
    m = 7
    lines = ["Datum;Region;Betrag Rp."]
    for region, amt in rows(m):
        lines.append(f"{LAST_DAY[m - 1]:02d}.{m:02d}.2025;{region};{int(round(amt * 100))}")
    write_cp1252(
        "2025-07.csv",
        "\n".join(lines) + "\n",
        "THE UNIT TRAP: integer Rappen, 100x out if bound naively",
    )


def build_two_betrag():
    """Net and gross, both literally named Betrag. The binder must refuse."""
    m = 8
    lines = ["Datum;Region;Betrag;Betrag"]
    for region, amt in rows(m):
        lines.append(
            f"{LAST_DAY[m - 1]:02d}.{m:02d}.2025;{region};{swiss(amt)};{swiss(amt * 1.081)}"
        )
    write_cp1252(
        "2025-08.csv",
        "\n".join(lines) + "\n",
        "THE AMBIGUITY TRAP: two columns named Betrag (net, gross)",
    )


def build_partial():
    """No region column. No plan reaches the target."""
    m = 11
    lines = ["Datum;Betrag"]
    for _, amt in rows(m):
        lines.append(f"{LAST_DAY[m - 1]:02d}.{m:02d}.2025;{swiss(amt)}")
    write_cp1252(
        "2025-11.csv",
        "\n".join(lines) + "\n",
        "THE PARTIAL EXPORT: no region column; must be a hard error",
    )


def build_extra_column():
    """An undeclared column, which the projection drops."""
    m = 12
    lines = ["Datum;Region;Betrag;Kundennummer"]
    for i, (region, amt) in enumerate(rows(m)):
        lines.append(
            f"{LAST_DAY[m - 1]:02d}.{m:02d}.2025;{region};{swiss(amt)};K{100000 + i:06d}"
        )
    write_cp1252(
        "2025-12.csv",
        "\n".join(lines) + "\n",
        "harmless extra column; must fit, Kundennummer dropped",
    )


MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def repack(path):
    """Pin zip stamps and dcterms:modified; see 09_legacy_formats.py."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        entries = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    entries = [
        (n, MODIFIED_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", d)
         if n == "docProps/core.xml" else d)
        for n, d in entries
    ]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for n, d in entries:
            info = zipfile.ZipInfo(n, date_time=ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            zout.writestr(info, d)
    os.replace(tmp, path)


def build_workbooks():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # --- 2025-09: a title row and a merged band above the real header -------
    m = 9
    wb = Workbook()
    ws = wb.active
    ws.title = "Umsatz"
    ws.append(["Muster AG — Monatsauswertung"])
    ws.append([None, None, "Umsatz 2025"])
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=3)
    ws.append(["Datum", "Region", "Betrag CHF"])
    for region, amt in rows(m):
        ws.append([f"{LAST_DAY[m - 1]:02d}.{m:02d}.2025", region, swiss(amt)])
    wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
    p = os.path.join(OUT, "2025-09.xlsx")
    wb.save(p)
    repack(p)
    note(p, "structural: title row + band above the header, `Betrag CHF`")
    _ = get_column_letter  # imported for clarity about the merged band's extent

    # --- 2025-10: English locale, ISO dates, an extra column ---------------
    m = 10
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Date", "Region", "Amount", "Discount"])
    for i, (region, amt) in enumerate(rows(m)):
        ws.append([f"2025-{m:02d}-{LAST_DAY[m - 1]:02d}", region, f"{amt:.2f}", f"{i * 5}"])
    wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
    p = os.path.join(OUT, "2025-10.xlsx")
    wb.save(p)
    repack(p)
    note(p, "English locale: Date/Amount, ISO dates, extra Discount")


TARGET_SQL = """-- The clean dataset we want out of a year of drifting monthly exports.
-- Hand-written, reviewed in git, versioned beside the data.
--
-- What a target may say is exactly what reaches the Arrow schema: a name, a
-- type, a nullability. Note what is absent — there is no date format here,
-- because that is a property of a file: these twelve exports carry two
-- different ones and both must land on this one DATE column.
--
-- `matches` is the other half. A target names what we *want*; the files are
-- somebody else's exports and say Datum, Betrag, Amount. Nothing bridges that
-- automatically, and a planner guessing at synonyms is exactly what this tool
-- does not do — so the synonyms are declared here, in the open, in a diff.

CREATE TABLE sales (
  month      DATE          NOT NULL OPTIONS(matches = 'Datum, Date, Buchungsdatum'),
  region     TEXT          NOT NULL OPTIONS(matches = 'Region, Kanton, Gebiet'),
  amount_chf DECIMAL(14,2) NOT NULL OPTIONS(matches = 'Betrag, Betrag CHF, Amount, Umsatz')
)
WITH (
  files      = '2025-*.csv, 2025-*.xlsx',
  date_order = 'dmy'
);
"""


# The same dataset with the three unfittable months excluded, so there is a
# target in the tree that actually resolves to a lock and a queryable relation.
# Excluding a file is one of the three *declared* softenings: it is written
# down, versioned, and visible in a diff — not a flag someone passed once.
SALES_OK_SQL = TARGET_SQL.replace(
    "CREATE TABLE sales (", "CREATE TABLE sales_ok ("
).replace(
    "  files      = '2025-*.csv, 2025-*.xlsx',",
    "  files      = '2025-*.csv, 2025-*.xlsx',\n"
    "  -- 07 is in Rappen, 08 has two columns called Betrag, 11 has no region.\n"
    "  exclude    = '2025-07.csv, 2025-08.csv, 2025-11.csv',",
)


def build_target():
    p = os.path.join(OUT, "sales.tdy.sql")
    with open(p, "w", encoding="utf-8") as f:
        f.write(TARGET_SQL)
    note(p, "the declared target; three of the twelve cannot reach it")

    p = os.path.join(OUT, "sales_ok.tdy.sql")
    with open(p, "w", encoding="utf-8") as f:
        f.write(SALES_OK_SQL)
    note(p, "the same, with the three unfittable months excluded")


def main():
    os.makedirs(OUT, exist_ok=True)
    build_ordinary_months()
    build_rappen()
    build_two_betrag()
    build_workbooks()
    build_partial()
    build_extra_column()
    build_target()

    fittable = [1, 2, 3, 4, 5, 6, 9, 10, 12]
    total = sum(amount(m, r) for m in fittable for r in range(4))
    print(f"\nground truth: {len(fittable)} fittable months, sum(amount_chf) = {total:.2f}")
    print("             2025-07 (Rappen), 2025-08 (ambiguous), 2025-11 (no region) must not join")


if __name__ == "__main__":
    main()
