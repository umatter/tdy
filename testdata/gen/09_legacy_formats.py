#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the `legacy_formats_*` fixture family for tdy (job key: legacy-formats).

Run from the repo root:  python3 testdata/gen/09_legacy_formats.py

Deterministic and idempotent: no randomness, and every zip entry is written
with a pinned timestamp, so re-running produces identical bytes and does not
silently stale a committed sidecar's blake3 fingerprint. (xlwt embeds no
wall clock, so the .xls files are byte-stable as written.)

WHY THIS FAMILY EXISTS
--------------------------------------------------------------------------
`src/sample.rs::guess_format` routes five extensions to FormatGuess::Excel:
xlsx, xls, xlsb, xlsm and ods. Only .xlsx had a fixture, so four fifths of
that claim rested on calamine's reputation rather than on a test. These are
three *different readers* inside calamine — the OOXML one (xlsx/xlsm), the
BIFF/CFB one (xls), and the OpenDocument one (ods) — reached through one
line of tdy code, and they are exactly the kind of thing that fails
silently: a misread cell is a wrong number, not a crash.

Not covered: .xlsb. It is a binary OOXML variant with no pure-Python
writer, so a fixture would mean committing a file this repo cannot
regenerate — which the "fixtures are generated, never hand-edited" rule
forbids. `guess_format` still routes it, and calamine still reads it; that
path is untested here and README says so.

DEPENDENCIES
--------------------------------------------------------------------------
openpyxl (xlsx/xlsm) and xlwt (xls). ODS is written with stdlib zipfile —
an .ods is a zip of XML, and hand-writing it avoids a third dependency.
If xlwt is missing this generator writes everything else and exits 0 with a
loud notice, because a missing legacy writer must not take the whole
fixture suite down; the committed .xls files stay valid either way.

GROUND TRUTH  (measured against target/release/tdy, not predicted)
--------------------------------------------------------------------------
  * calamine renders a `datetime.date` cell (xlsx and xls) and an ODS
    office:value-type="date" cell identically, as "YYYY-MM-DD" — so tier 1
    types all three as date with format "%Y-%m-%d". This differs from a
    `datetime.datetime` cell, which renders as "YYYY-MM-DD HH:MM:SS" (see
    02_excel_nightmares.py, whose fixtures use datetimes).
  * ODS `table:number-columns-repeated` and `table:number-rows-repeated`
    are expanded by calamine, so a sparse row keeps its column alignment.
    Fixture 2 pins that: if it ever regresses, values slide left into the
    wrong columns and every one of them is silently wrong.
  * BIFF8 stores strings as UTF-16LE while xlsx and ods store UTF-8 XML.
    Fixture 1 carries "Zürich"/"Genève" through all four containers.

--------------------------------------------------------------------------
FIXTURES  (all in testdata/, all named legacy_formats_*)
--------------------------------------------------------------------------

1. legacy_formats_same_table.{xlsx,xlsm,xls,ods}
   ONE logical table in four containers. A one-row title, a blank row, a
   header row, three body rows:

       Filialumsatz 2025
       (blank)
       Filiale | Eroeffnet  | Umsatz  | Mitarbeitende
       Zürich  | 2025-03-04 | 1234.50 | 12
       Genève  | 2025-06-15 |  987.25 |  8
       Lugano  | 2025-09-01 | 2500.00 | 15

   Stresses the routing claim itself. The invariant is not "each parses"
   but "all four parse the *same*": same transforms, same dtypes, same
   values. A container that quietly drops the accented characters, shifts a
   column or types Umsatz as float instead of decimal breaks the equality
   even though it parses fine on its own.
   Correct parse: skip_rows head = 2 + promote_header rows = 1 →
   3 rows x 4 columns [filiale Utf8, eroeffnet Date("%Y-%m-%d"),
   umsatz Decimal(38,2), mitarbeitende Int64];
   sum(umsatz) = 4721.75; sum(mitarbeitende) = 35;
   filiale values are exactly ["Zürich", "Genève", "Lugano"].

2. legacy_formats_ods_sparse.ods
   The ODS-specific trap, in the form a real LibreOffice export produces:
   row 2 writes A and E with a `number-columns-repeated="3"` empty run
   between them, and the last two rows are a single element carrying
   `number-rows-repeated="2"`. Nothing here is exotic — LibreOffice emits
   both attributes for any sheet with gaps or duplicate rows.
   Correct parse: promote_header rows = 1 → 3 rows x 5 columns
   [a..e]; row 0 = (1, NULL, NULL, NULL, 5); rows 1 and 2 are both
   (1, 2, 3, 4, 5). The point is column e: it must hold 5, not NULL, and
   must not have slid to column b.
   Measured tier-1 behaviour: the mostly-empty second row reads as more
   title block, so the sniffer skips two leading rows, finds no header and
   emits col_1..col_5 over the two dense rows — confidence 0.65, with
   "skipped 2 leading row(s)" and "no header row detected" in notes. That
   is the honest outcome, not a defect: it loses a row rather than
   inventing one, and says so loudly enough to escalate. The correct parse
   above needs the hand-written spec that tests/formats.rs pins.

3. legacy_formats_xls_cover_sheet.xls
   Sheet selection on the BIFF reader. Sheets are ["Deckblatt", "Daten"]:
   sheet 0 is a prose cover page, the data is on sheet 1. Proves the legacy
   reader exposes a sheet *list* the selector can score, not just sheet 0.
   Correct parse: extraction sheet_name = "Daten" + promote_header rows = 1
   → 4 rows x 3 columns [beleg Int64, betrag Decimal(38,2), waehrung Utf8];
   sum(betrag) = 6666.00; 1 row has waehrung = 'EUR'.
   Measured tier-1 behaviour: picks "Daten" unaided, confidence 0.90 — the
   selector scores sheets by numeric content first, and this cover page has
   none. Its .xlsx twin (excel_nightmares_cover_sheet.xlsx) is the harder
   case and is still mis-picked: there, sheet 1 is a two-column glossary
   that *looks* tabular. Keep both; they bracket the selector rather than
   duplicating each other.
"""

import os
import re
import sys
import zipfile
from datetime import date, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
OUT = os.path.join(REPO, "testdata")
PREFIX = "legacy_formats_"

# Pinned so the zip containers are byte-stable across regenerations.
ZIP_EPOCH = (2026, 1, 1, 0, 0, 0)

TITLE = "Filialumsatz 2025"
HEADER = ["Filiale", "Eroeffnet", "Umsatz", "Mitarbeitende"]
BODY = [
    ["Zürich", date(2025, 3, 4), 1234.50, 12],
    ["Genève", date(2025, 6, 15), 987.25, 8],
    ["Lugano", date(2025, 9, 1), 2500.00, 15],
]


def note(path, what):
    print(f"wrote {os.path.relpath(path, REPO)} ({os.path.getsize(path)} bytes) - {what}")


# ---------------------------------------------------------------------------
# ODS: an OpenDocument spreadsheet is a zip of XML, so stdlib is enough.
# ---------------------------------------------------------------------------

ODS_NS = (
    'xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" '
    'xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0" '
    'xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"'
)

# calamine refuses an .ods without a manifest, so it is not optional padding.
ODS_MANIFEST = (
    '<?xml version="1.0" encoding="UTF-8"?>'
    '<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0"'
    ' manifest:version="1.2">'
    '<manifest:file-entry manifest:full-path="/"'
    ' manifest:media-type="application/vnd.oasis.opendocument.spreadsheet"/>'
    '<manifest:file-entry manifest:full-path="content.xml" manifest:media-type="text/xml"/>'
    "</manifest:manifest>"
)


def xml_escape(s):
    return (
        str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )


def ods_cell(v):
    """One <table:table-cell>. Cells are typed, which is what calamine reads."""
    if v is None:
        return "<table:table-cell/>"
    if isinstance(v, date):
        d = v.isoformat()
        return (
            '<table:table-cell office:value-type="date" office:date-value="%s">'
            "<text:p>%s</text:p></table:table-cell>" % (d, d)
        )
    if isinstance(v, bool):
        return (
            '<table:table-cell office:value-type="boolean" office:boolean-value="%s">'
            "<text:p>%s</text:p></table:table-cell>"
            % (str(v).lower(), str(v).lower())
        )
    if isinstance(v, (int, float)):
        return (
            '<table:table-cell office:value-type="float" office:value="%s">'
            "<text:p>%s</text:p></table:table-cell>" % (v, v)
        )
    t = xml_escape(v)
    return (
        '<table:table-cell office:value-type="string"><text:p>%s</text:p>'
        "</table:table-cell>" % t
    )


def ods_row(cells, repeat=None):
    attr = ' table:number-rows-repeated="%d"' % repeat if repeat else ""
    return "<table:table-row%s>%s</table:table-row>" % (attr, "".join(cells))


def write_ods(path, sheets):
    """`sheets` maps a sheet name to a list of pre-rendered row strings."""
    body = "".join(
        '<table:table table:name="%s">%s</table:table>' % (xml_escape(name), "".join(rows))
        for name, rows in sheets.items()
    )
    content = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<office:document-content %s office:version="1.2">'
        "<office:body><office:spreadsheet>%s</office:spreadsheet>"
        "</office:body></office:document-content>" % (ODS_NS, body)
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as z:
        # The mimetype entry must come first and be stored uncompressed; that
        # is how a reader identifies the package before unzipping it.
        mt = zipfile.ZipInfo("mimetype", date_time=ZIP_EPOCH)
        mt.compress_type = zipfile.ZIP_STORED
        mt.create_system = 0
        mt.external_attr = 0o644 << 16
        z.writestr(mt, "application/vnd.oasis.opendocument.spreadsheet")
        for name, data in (
            ("META-INF/manifest.xml", ODS_MANIFEST),
            ("content.xml", content),
        ):
            info = zipfile.ZipInfo(name, date_time=ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            z.writestr(info, data.encode("utf-8"))


# ---------------------------------------------------------------------------
# xlsx / xlsm, via openpyxl.
# ---------------------------------------------------------------------------


MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def repack_deterministic(path):
    """Pin zip entry timestamps and dcterms:modified.

    openpyxl stamps every zip entry with the wall clock, and rewrites
    dcterms:modified at save time whatever `wb.properties` says — so pinning
    the property is not enough on its own and the XML has to be patched here.
    """
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        entries = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    entries = [
        (name, MODIFIED_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", data)
         if name == "docProps/core.xml" else data)
        for name, data in entries
    ]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for name, data in entries:
            info = zipfile.ZipInfo(name, date_time=ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            zout.writestr(info, data)
    os.replace(tmp, path)


def build_ooxml():
    from openpyxl import Workbook

    for ext in ("xlsx", "xlsm"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Umsatz"
        # openpyxl writes dcterms:created/modified from the wall clock, which
        # would change the bytes on every regeneration and stale every
        # committed sidecar's fingerprint. Pinning them here means
        # repack_deterministic only has to deal with the zip metadata.
        wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
        ws.append([TITLE])
        ws.append([])
        ws.append(HEADER)
        for r in BODY:
            ws.append(r)
        path = os.path.join(OUT, PREFIX + "same_table." + ext)
        wb.save(path)
        repack_deterministic(path)
        note(path, "the shared table, OOXML container")


def build_ods():
    rows = [ods_row([ods_cell(TITLE)]), ods_row([]), ods_row([ods_cell(c) for c in HEADER])]
    rows += [ods_row([ods_cell(c) for c in r]) for r in BODY]
    path = os.path.join(OUT, PREFIX + "same_table.ods")
    write_ods(path, {"Umsatz": rows})
    note(path, "the shared table, OpenDocument container")

    # Fixture 2: the repeated-cell trap.
    head = ods_row([ods_cell(c) for c in "abcde"])
    sparse = ods_row(
        [
            ods_cell(1),
            '<table:table-cell table:number-columns-repeated="3"/>',
            ods_cell(5),
        ]
    )
    dup = ods_row([ods_cell(i) for i in range(1, 6)], repeat=2)
    path = os.path.join(OUT, PREFIX + "ods_sparse.ods")
    write_ods(path, {"Sparse": [head, sparse, dup]})
    note(path, "number-columns-repeated / number-rows-repeated")


# ---------------------------------------------------------------------------
# xls (BIFF8), via xlwt. Optional: see the module docstring.
# ---------------------------------------------------------------------------


def build_xls():
    try:
        import xlwt
    except ImportError:
        print(
            "NOTE: xlwt is not installed - skipping the .xls fixtures.\n"
            "      The committed files stay valid; install it (pip install xlwt)\n"
            "      to regenerate them.",
            file=sys.stderr,
        )
        return

    datestyle = xlwt.easyxf(num_format_str="YYYY-MM-DD")

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Umsatz")
    ws.write(0, 0, TITLE)
    # Row 1 stays blank.
    for c, h in enumerate(HEADER):
        ws.write(2, c, h)
    for r, row in enumerate(BODY, start=3):
        for c, v in enumerate(row):
            ws.write(r, c, v, datestyle if isinstance(v, date) else xlwt.Style.default_style)
    path = os.path.join(OUT, PREFIX + "same_table.xls")
    wb.save(path)
    note(path, "the shared table, BIFF8 container")

    # Fixture 3: a decoy cover sheet ahead of the data, on the legacy reader.
    wb = xlwt.Workbook(encoding="utf-8")
    cover = wb.add_sheet("Deckblatt")
    for r, line in enumerate(
        [
            "Muster AG",
            "Belegjournal, Abteilung Controlling",
            "Alle Betraege in CHF sofern nicht anders vermerkt.",
            "Fragen an controlling@muster.example",
        ]
    ):
        cover.write(r, 0, line)
    data = wb.add_sheet("Daten")
    for c, h in enumerate(["Beleg", "Betrag", "Waehrung"]):
        data.write(0, c, h)
    for r, row in enumerate(
        [
            [4001, 1234.50, "CHF"],
            [4002, 2500.00, "CHF"],
            [4003, 987.25, "EUR"],
            [4004, 1944.25, "CHF"],
        ],
        start=1,
    ):
        for c, v in enumerate(row):
            data.write(r, c, v)
    path = os.path.join(OUT, PREFIX + "xls_cover_sheet.xls")
    wb.save(path)
    note(path, "decoy cover sheet, data on sheet 1")


def main():
    os.makedirs(OUT, exist_ok=True)
    build_ooxml()
    build_ods()
    build_xls()


if __name__ == "__main__":
    main()
