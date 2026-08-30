#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the `declared_size_*` fixtures for tdy (job key: declared-size).

Run from the repo root:  python3 testdata/gen/10_declared_size.py

Deterministic and idempotent; openpyxl for the .xlsx, stdlib for the .ods.

WHY THIS FAMILY EXISTS
--------------------------------------------------------------------------
Every other limit in tdy is applied to a table that already exists, which is
fine for delimited text: the file has to be as big as the data. A spreadsheet
*declares* its geometry, so a few hundred bytes can ask for tens of
gigabytes, and the request is granted before any limit is consulted.

Measured before the guard existed (`src/xlguard.rs`):

    898-byte .ods declaring 40M cells  ->  4.78 GB peak RSS, 7.6 s
    4.8 KB .xlsx with one far cell     ->  2.11 GB, and SIGABRT under a
                                           3 GB cap: "memory allocation of
                                           1280000640 bytes failed"

An abort is the one failure mode tdy is not allowed to have — not a loud
error naming the problem, but the process dying. These fixtures are the
regression: each must be refused in milliseconds with a sentence.

The control file matters as much as the bombs. LibreOffice pads every sheet
it writes out to the full 1,048,576-row grid with valueless cells, so a
guard that counted declared cells naively would refuse almost every .ods in
existence. `declared_size_ods_padded_like_libreoffice.ods` is an ordinary
three-row table wearing that padding, and it must parse normally.

--------------------------------------------------------------------------
FIXTURES  (all in testdata/, all named declared_size_*)
--------------------------------------------------------------------------

1. declared_size_ods_declared_grid.ods  (~900 bytes)
   One `<table:table-row table:number-rows-repeated="50000000">` carrying 20
   valued cells: 1,000,000,020 cells claimed from 900 bytes on disk, an
   amplification of about 1.1 million to one. Semantic expansion, not
   compression — content.xml is *tiny*, so a zip-ratio check would see
   nothing wrong.
   Correct behaviour: refused by `xlguard::preflight` before the workbook is
   opened at all (calamine's Ods reader parses content.xml eagerly, so
   opening it is already too late). Error names the declared count and the
   limit, and says which knob raises it.

2. declared_size_xlsx_phantom_grid.xlsx  (~4.8 KB)
   A1 holds a value and so does row 1,000,000 column 100 — nothing in
   between. The used range is therefore 100,000,000 cells. This is the
   ordinary shape of a spreadsheet someone has scrolled to the bottom of and
   typed in, not an attack, which is why it must fail politely.
   Correct behaviour: refused by `engine::checked_worksheet_range`, which
   reads `XlsxCellReader::dimensions()` — the `<dimension>` the file
   declares — before the grid is built.

3. declared_size_ods_padded_like_libreoffice.ods  (~1 KB)
   THE CONTROL. Three real rows of data, then
   `number-rows-repeated="1048573"` over a `number-columns-repeated="1024"`
   valueless cell, exactly as LibreOffice writes it. Declares over a billion
   cell positions; contains three rows of data.
   Correct parse: promote_header rows = 1 -> 3 rows x 3 columns
   [stadt Utf8, jahr Int64, betrag Decimal(38,2)];
   sum(betrag) = 6041.75; stadt = ["Bern", "Chur", "Sion"].
   If this file ever starts erroring, the guard has become useless in
   practice however well it stops the bombs.
"""

import os
import re
import zipfile
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
OUT = os.path.join(REPO, "testdata")
PREFIX = "declared_size_"

ZIP_EPOCH = (2026, 1, 1, 0, 0, 0)

ODS_NS = (
    'xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" '
    'xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0" '
    'xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"'
)

ODS_MANIFEST = (
    '<?xml version="1.0" encoding="UTF-8"?>'
    '<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0"'
    ' manifest:version="1.2">'
    '<manifest:file-entry manifest:full-path="/"'
    ' manifest:media-type="application/vnd.oasis.opendocument.spreadsheet"/>'
    '<manifest:file-entry manifest:full-path="content.xml" manifest:media-type="text/xml"/>'
    "</manifest:manifest>"
)


def note(path, what):
    print(f"wrote {os.path.relpath(path, REPO)} ({os.path.getsize(path)} bytes) - {what}")


def value_cell(v, kind="float"):
    if kind == "string":
        return (
            '<table:table-cell office:value-type="string"><text:p>%s</text:p>'
            "</table:table-cell>" % v
        )
    return (
        '<table:table-cell office:value-type="float" office:value="%s">'
        "<text:p>%s</text:p></table:table-cell>" % (v, v)
    )


def write_ods(name, table_name, rows_xml, what):
    content = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<office:document-content %s office:version="1.2">'
        "<office:body><office:spreadsheet>"
        '<table:table table:name="%s">%s</table:table>'
        "</office:spreadsheet></office:body></office:document-content>"
        % (ODS_NS, table_name, rows_xml)
    )
    path = os.path.join(OUT, PREFIX + name)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as z:
        mt = zipfile.ZipInfo("mimetype", date_time=ZIP_EPOCH)
        mt.compress_type = zipfile.ZIP_STORED
        mt.create_system = 0
        mt.external_attr = 0o644 << 16
        z.writestr(mt, "application/vnd.oasis.opendocument.spreadsheet")
        for entry, data in (
            ("META-INF/manifest.xml", ODS_MANIFEST),
            ("content.xml", content),
        ):
            info = zipfile.ZipInfo(entry, date_time=ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            z.writestr(info, data.encode("utf-8"))
    note(path, what)


def build_ods_bomb():
    header = "<table:table-row>%s</table:table-row>" % "".join(
        value_cell("c%d" % i, "string") for i in range(20)
    )
    body = '<table:table-row table:number-rows-repeated="50000000">%s</table:table-row>' % (
        value_cell("1") * 20
    )
    write_ods(
        "ods_declared_grid.ods",
        "Bomb",
        header + body,
        "1,000,000,020 cells declared from ~900 bytes",
    )


def build_ods_control():
    """An ordinary table wearing LibreOffice's full-grid padding."""
    head = "<table:table-row>%s</table:table-row>" % "".join(
        value_cell(h, "string") for h in ("Stadt", "Jahr", "Betrag")
    )
    body = "".join(
        "<table:table-row>%s%s%s</table:table-row>"
        % (value_cell(stadt, "string"), value_cell(jahr), value_cell(betrag))
        for stadt, jahr, betrag in (
            ("Bern", 2025, "1234.50"),
            ("Chur", 2025, "987.25"),
            ("Sion", 2025, "3820.00"),
        )
    )
    # Exactly the shape LibreOffice emits: the rest of the grid, valueless.
    pad = (
        '<table:table-row table:number-rows-repeated="1048573">'
        '<table:table-cell table:number-columns-repeated="1024"/>'
        "</table:table-row>"
    )
    write_ods(
        "ods_padded_like_libreoffice.ods",
        "Umsatz",
        head + body + pad,
        "THE CONTROL: full-grid padding, must still parse",
    )


def build_xlsx_phantom():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "wert"
    ws["A2"] = 1
    # One cell in the far corner is all it takes to declare the whole grid.
    ws.cell(row=1_000_000, column=100, value=1)
    path = os.path.join(OUT, PREFIX + "xlsx_phantom_grid.xlsx")
    wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
    wb.save(path)
    _repack(path)
    note(path, "100,000,000-cell used range from ~4.8 KB")


_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _repack(path):
    """Pin zip stamps and dcterms:modified; see 09_legacy_formats.py."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        entries = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    entries = [
        (n, _MODIFIED_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", d)
         if n == "docProps/core.xml" else d)
        for n, d in entries
    ]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for n, d in entries:
            info = zipfile.ZipInfo(n, date_time=ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            zout.writestr(info, d)
    os.replace(tmp, path)


def main():
    os.makedirs(OUT, exist_ok=True)
    build_ods_bomb()
    build_ods_control()
    build_xlsx_phantom()


if __name__ == "__main__":
    main()
