#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the `sheet_frames_*` fixtures for tdy (job key: sheet-frames).

Run from the repo root:  python3 testdata/gen/14_sheet_frames.py

Deterministic and idempotent; needs openpyxl.

WHY THIS FAMILY EXISTS
--------------------------------------------------------------------------
The sheet is Excel's version of the JSON record-array problem: a workbook
holds a cover page, a legend, and somewhere in the middle the actual data,
and the sniffer's pick ("the biggest sheet with numbers in it") is a
ranking, not a proof. A declared table turns it into one — `tdy fit` frames
EVERY sheet and tries the declaration against each:

    exactly one fits  ->  proved by elimination
    several fit       ->  refused with every sheet named; two well-typed
                          answers with different totals is a guess
    none fit          ->  the ordinary gap report, for the ranked sheet

Each sheet is framed independently (title rows, merged bands, footers are
facts about a sheet, not a file), which is what `sniff_excel_sheet` exists
for.

--------------------------------------------------------------------------
FIXTURES  (all in testdata/, named sheet_frames_*)
--------------------------------------------------------------------------

1. sheet_frames_one_fits.xlsx
   Three sheets: "Hinweise" (a prose cover page), "Daten" (a title row above
   Datum/Region/Betrag data — its OWN framing problem, solved per sheet),
   "Legende" (a two-column glossary). Only "Daten" fits a target declaring
   (month DATE, region TEXT, amount DECIMAL). "Daten" is deliberately NOT
   the biggest sheet, so the ranked pick alone would not settle it.
   Ground truth: sum(amount) = 1'090.00 over 4 rows.

2. sheet_frames_two_fit.xlsx
   "Q1" and "Q2": the same shape, different numbers (sums 600.00 and
   1'500.00). Both fit; must be refused with both sheets named.
"""

import os
import re
import zipfile
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
OUT = os.path.join(REPO, "testdata")
ZIP_EPOCH = (2026, 1, 1, 0, 0, 0)

MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def repack(path):
    """Pin zip stamps and dcterms:modified; see 09_legacy_formats.py."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        entries = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    entries = [
        (n, MODIFIED_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", d)
         if n == "docProps/core.xml" else d)
        for n, d in entries
    ]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for n, d in entries:
            info = zipfile.ZipInfo(n, date_time=ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0o644 << 16
            zout.writestr(info, d)
    os.replace(tmp, path)


def note(path, what):
    print(f"wrote {os.path.relpath(path, REPO)} ({os.path.getsize(path)} bytes) - {what}")


def save(wb, name, what):
    wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
    p = os.path.join(OUT, "sheet_frames_" + name)
    wb.save(p)
    repack(p)
    note(p, what)


def build_one_fits():
    from openpyxl import Workbook

    wb = Workbook()

    # A wordy cover page — bigger than the data sheet, all prose. It exists
    # to make "biggest sheet" the wrong heuristic.
    ws = wb.active
    ws.title = "Hinweise"
    ws.append(["Monatsauswertung Muster AG"])
    ws.append([])
    for i in range(1, 13):
        ws.append([f"Hinweis {i}", "Diese Auswertung wird monatlich erstellt und",
                   "ersetzt keine revidierte Jahresrechnung."])

    ws = wb.create_sheet("Daten")
    ws.append(["Muster AG — internes Reporting"])   # a title row: per-sheet framing
    ws.append(["Datum", "Region", "Betrag"])
    for day, region, amount in [
        ("05.08.2025", "Ost", "250.00"),
        ("12.08.2025", "West", "260.00"),
        ("19.08.2025", "Nord", "270.00"),
        ("26.08.2025", "Sued", "310.00"),
    ]:
        ws.append([day, region, amount])

    ws = wb.create_sheet("Legende")
    ws.append(["Kürzel", "Bedeutung"])
    for k, v in [("Ost", "Ostschweiz"), ("West", "Romandie"),
                 ("Nord", "Basel und Umgebung"), ("Sued", "Tessin")]:
        ws.append([k, v])

    save(wb, "one_fits.xlsx", "only 'Daten' fits (Datum, Region, Betrag); sum 1090.00")


def build_two_fit():
    from openpyxl import Workbook

    wb = Workbook()
    for title, base in [("Q1", 190), ("Q2", 490)]:
        ws = wb.active if title == "Q1" else wb.create_sheet(title)
        ws.title = title
        ws.append(["Datum", "Region", "Betrag"])
        for i, (m, region) in enumerate([("01", "Ost"), ("02", "West"), ("03", "Nord")]):
            ws.append([f"28.{m}.2025", region, f"{base + 10 * i}.00"])
    save(wb, "two_fit.xlsx", "Q1 (600.00) and Q2 (1500.00) BOTH fit: must be refused")


def main():
    os.makedirs(OUT, exist_ok=True)
    build_one_fits()
    build_two_fit()
    print("\nground truth: Daten = 1090.00 over 4 rows; Q1 = 600.00, Q2 = 1500.00")


if __name__ == "__main__":
    main()
